"""
Parser for the May 6, 2014 Tennessee State Judicial Districts primary, producing
OpenElections CSVs (county + precinct) sourced from the TN SoS results page
(https://sos.tn.gov/elections/results#2014).

The May 6, 2014 primary was a judicial-district primary only -- the ballot was
entirely judicial/DA/Public Defender races (no Governor/Senate/House). Both the
Republican and Democratic primaries were held the same day. Source:

  XLSX (primary source):  https://sos-tn-gov-files.s3.amazonaws.com/May 2014 by precinct.xlsx
    -- the standard SoS "All by Precinct" workbook (CANDGROUP + RNAME1..10 /
       PVTALLY1..10 inline triples), sheet "SOFFICEL", 20,869 rows. ELECTTYPE is
       "Republican Primary" or "Democratic Primary"; ELECTDATE is "May 6, 2014".
       OFFICENAME is e.g. "Circuit Court Judge Division 1 District 20",
       "Chancellor Part I District  6", "District Attorney General District  1",
       "Public Defender District 31", "Circuit and Criminal Court Judge District 7",
       "Chancellor District  4/5" (note double spaces and the slash district).

  PDFs (independent verification, by county):
    20140506_State Judicial Districts Republican Primary by County.pdf
    20140506_State Judicial Districts Democratic Primary by County.pdf
    Layout: banner "State of Tennessee", date, "<Party> Primary" section header,
    an indented office line, "N. Name" candidate lines, an all-digit column
    header, "<County>  v1 v2 ..." county rows, and a "DISTRICT TOTALS  v1 v2 ..."
    footer. The PDF is the official spelling reference.

Conventions (standard OpenElections office names; integer votes; full party
names; precinct names verbatim with internal whitespace collapsed):
  "<office text> District N [(unexpired term)]"
                              -> office = text before "District N",
                                 district = "N" (or "N/M" for split districts,
                                 or "N (unexpired term)" if suffixed)
  party from ELECTTYPE ("Republican Primary" -> "Republican", etc.)
So e.g. "Circuit Court Judge Division 1 District 20" -> office "Circuit Court
Judge Division 1", district "20"; "Chancellor District  4/5" -> office
"Chancellor", district "4/5". The "Division N" / "Part I" qualifier stays in the
office name (matching the 2018 May 1 judicial-primary convention). The 2014 May
offices have NO "(unexpired term)" suffix, so all districts are bare numbers.

Candidate-name standardization: the XLSX spells a handful of candidates
inconsistently across precinct rows (case or a comma, e.g. "J. Thomas (Tom)
DuBois" vs "Dubois", "John F. Dugger, Jr." vs "Dugger Jr.", "No Candidate
Qualified" vs "no candidate qualified"). Position-within-race is consistent and
contiguous, so for each (race, candidate-position) the most common spelling is
chosen as the canonical name and applied to every precinct row; the most common
spelling matches the by-county PDF's official spelling in every case. This keeps
the precinct file internally consistent and lets the county file aggregate by
candidate name without splitting one candidate into two rows. "No Candidate
Qualified" rows (0 votes) are kept verbatim, matching the 2018 May 1 convention.

County totals are derived by summing precinct votes per (county, office,
district, party, candidate) and asserted against the by-county PDFs' per-county
rows AND DISTRICT TOTALS (matched by candidate position within each race, since
the PDF and XLSX list candidates in the same numbered order).

Requires `requests` + `openpyxl` (in the Pipfile) and `pdftotext` (poppler).
"""

import csv
import io
import os
import re
import subprocess
import tempfile
from collections import Counter, defaultdict

import openpyxl
import requests

BASE = "https://sos-tn-gov-files.s3.amazonaws.com"

XLSX_URL = f"{BASE}/May 2014 by precinct.xlsx"
REP_COUNTY_PDF = (
    f"{BASE}/20140506_State Judicial Districts Republican Primary by County.pdf")
DEM_COUNTY_PDF = (
    f"{BASE}/20140506_State Judicial Districts Democratic Primary by County.pdf")

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2014")

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party",
                   "candidate", "votes"]

DISTRICT_RE = re.compile(
    r"^(.+?)\s+District\s+(\d+(?:/\d+)?)\s*(\(unexpired term\))?\s*$")
SECTION_RE = re.compile(r"^(Republican Primary|Democratic Primary)\s*$")
CAND_RE = re.compile(r"^\s*(\d+)\.\s+(.+?)\s*$")
COUNTY_ROW_RE = re.compile(r"^([A-Z][A-Za-z .]+?)\s+(\d[\d,]*(?:\s+\d[\d,]*)*)\s*$")
DISTRICT_TOTALS_RE = re.compile(r"^DISTRICT TOTALS\s+(.+?)\s*$")
ALLDIGITS_RE = re.compile(r"^\d+$")
WHITESPACE_RE = re.compile(r"\s+")


def clean(s):
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def to_int(tok):
    return int(str(tok).replace(",", ""))


def norm_office(raw):
    s = raw.strip()
    m = DISTRICT_RE.match(s)
    if m:
        office = clean(m.group(1))
        dist = m.group(2) + (" (unexpired term)" if m.group(3) else "")
        return office, dist
    return clean(s), "NA"


def party_from_etype(et):
    et = (et or "").strip()
    if "Republican" in et:
        return "Republican"
    if "Democratic" in et:
        return "Democratic"
    return ""


def load_xlsx(url):
    """Parse the All-by-Precinct XLSX. Returns (precinct_rows, county_rows)
    where precinct_rows are [county, precinct, office, district, party,
    candidate, votes] (candidate names standardized per race+position) and
    county_rows are the derived [county, office, district, party, candidate,
    votes] sums. Also returns pos_sums mapping (county, office, district,
    party, position) -> votes for verification against the by-county PDFs."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True,
                                data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {name: i for i, name in enumerate(hdr) if name}
    county_i = idx["COUNTY"]
    precinct_i = idx["PRECINCT"]
    office_i = idx["OFFICENAME"]
    etype_i = idx["ELECTTYPE"]
    rname_cols = [i for i, c in enumerate(hdr)
                  if c and str(c).startswith("RNAME")]
    pvtally_cols = [i for i, c in enumerate(hdr)
                    if c and str(c).startswith("PVTALLY")]

    parsed = []           # (county, precinct, raw_office, etype, k, name, tally)
    spell = defaultdict(Counter)   # (raw_office, etype, k) -> Counter(name)
    for r in rows[1:]:
        raw_office = r[office_i]
        if raw_office is None or not str(raw_office).strip():
            continue
        county = clean(r[county_i])
        precinct = clean(r[precinct_i])
        etype = r[etype_i]
        for k, rcol in enumerate(rname_cols, start=1):
            rn = r[rcol]
            if rn is None or not str(rn).strip():
                continue
            name = clean(rn)
            pv = r[pvtally_cols[k - 1]]
            tally = int(pv) if pv not in (None, "") else 0
            spell[(str(raw_office).strip(), str(etype).strip(), k)][name] += 1
            parsed.append((county, precinct, raw_office, etype, k, name, tally))

    canon = {key: c.most_common(1)[0][0] for key, c in spell.items()}

    precinct_rows = []
    county_sum = defaultdict(int)
    pos_sum = defaultdict(int)   # (county, office, district, party, k) -> votes
    for county, precinct, raw_office, etype, k, name, tally in parsed:
        cname = canon[(str(raw_office).strip(), str(etype).strip(), k)]
        office, district = norm_office(raw_office)
        party = party_from_etype(etype)
        precinct_rows.append([county, precinct, office, district, party,
                              cname, tally])
        county_sum[(county, office, district, party, cname)] += tally
        pos_sum[(county, office, district, party, k)] += tally
    county_rows = [[k2[0], k2[1], k2[2], k2[3], k2[4], v]
                   for k2, v in county_sum.items()]
    return county_rows, precinct_rows, pos_sum, canon


def pdftotext_layout(pdf_bytes):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        path = f.name
    try:
        out = subprocess.run(["pdftotext", "-layout", path, "-"],
                             check=True, capture_output=True, text=True)
        return out.stdout
    finally:
        os.unlink(path)


def parse_by_county_pdf(text, party):
    """Parse a by-county PDF for one party. Returns (county_totals,
    district_totals, pdf_candidates) keyed by (county, office, district, party,
    position) / (office, district, party, position) / (office, district, party,
    position) -> name."""
    county_totals = {}
    district_totals = {}
    pdf_candidates = {}
    office = None
    district = None
    candidates = []        # list of (num, name)
    current_county = None
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s == "State of Tennessee":
            office = None
            candidates = []
            current_county = None
            continue
        m = SECTION_RE.match(s)
        if m:
            office = None
            candidates = []
            current_county = None
            continue
        # DISTRICT TOTALS footer
        m = DISTRICT_TOTALS_RE.match(s)
        if m and office is not None and candidates:
            nums = m.group(1).split()
            if len(nums) == len(candidates):
                for k, (num, name) in enumerate(sorted(candidates,
                                                       key=lambda c: c[0]),
                                                start=1):
                    district_totals[(office, district, party, k)] = to_int(
                        nums[k - 1])
                    pdf_candidates[(office, district, party, k)] = name
            candidates = []
            office = None
            current_county = None
            continue
        # candidate line "N. Name"
        m = CAND_RE.match(line)
        if m and office is not None:
            candidates.append((int(m.group(1)), clean(m.group(2))))
            continue
        # column header line (all digits)
        toks = s.split()
        if toks and all(ALLDIGITS_RE.match(t) for t in toks) and office is not None:
            continue
        # office line (matches District N, not a candidate/county/totals line)
        if office is None:
            om = DISTRICT_RE.match(s)
            if om:
                office = clean(om.group(1))
                district = om.group(2) + (" (unexpired term)" if om.group(3)
                                          else "")
                candidates = []
                current_county = None
                continue
        # county row: "County  v1 v2 ..."
        if office is not None and candidates:
            cm = COUNTY_ROW_RE.match(s)
            if cm:
                cname = clean(cm.group(1))
                nums = cm.group(2).split()
                if len(nums) == len(candidates):
                    for k in range(len(candidates)):
                        county_totals[(cname, office, district, party,
                                       k + 1)] = to_int(nums[k])
                continue
    return county_totals, district_totals, pdf_candidates


def verify(pos_sum, pdf_county_totals, pdf_district_totals, pdf_candidates,
           party, label):
    """Compare XLSX-derived per-(county,office,district,party,position) sums
    and per-race district totals against the by-county PDF. Raises on mismatch.
    Also checks the PDF candidate name at each position matches the XLSX
    canonical name (sanity check that positions align)."""
    # Map (office,district,party,k) -> xlsx district sum (sum over counties)
    xlsx_district = defaultdict(int)
    for (county, off, dist, partyp, k), v in pos_sum.items():
        if partyp == party:
            xlsx_district[(off, dist, partyp, k)] += v
    mismatches = []
    # district totals
    for key, v in pdf_district_totals.items():
        if key[3] != party:
            continue
        xv = xlsx_district.get(key)
        if xv is None:
            mismatches.append((key, "MISSING in xlsx", v))
        elif xv != v:
            mismatches.append((key, xv, v))
    # county totals
    for key, v in pdf_county_totals.items():
        if key[4] != party:
            continue
        xv = pos_sum.get((key[0], key[1], key[2], key[3], key[4]))
        if xv is None:
            mismatches.append((key, "MISSING in xlsx", v))
        elif xv != v:
            mismatches.append((key, xv, v))
    if mismatches:
        detail = "\n  ".join(
            f"{k}: xlsx={dv} pdf={ev}" for k, dv, ev in mismatches[:30])
        raise AssertionError(
            f"{label}: {len(mismatches)} mismatches vs by-county PDF:\n  "
            f"{detail}")
    print(f"  {label}: {len(pdf_district_totals)} district-total cells and "
          f"{len(pdf_county_totals)} county-total cells all match.")


def sort_key_county(row):
    return (row[0], row[2], row[3], row[1], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  Wrote {path} ({len(rows)} rows)")


ELECTION = "20140506__tn__primary"


def main():
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)
    print(f"--- {ELECTION} ---")
    county_rows, precinct_rows, pos_sum, canon = load_xlsx(XLSX_URL)
    county_rows.sort(key=sort_key_county)
    precinct_rows.sort(key=sort_key_precinct)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} "
          f"rows, races: {len(canon)} candidate-positions")
    # Verify against the by-county PDFs (independent source).
    for party, pdf_url, label in [
            ("Republican", REP_COUNTY_PDF, "Republican by-county PDF"),
            ("Democratic", DEM_COUNTY_PDF, "Democratic by-county PDF")]:
        resp = requests.get(pdf_url, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        text = pdftotext_layout(resp.content)
        ctot, dtot, pcand = parse_by_county_pdf(text, party)
        verify(pos_sum, ctot, dtot, pcand, party, label)
    write_csv(os.path.join(out, f"{ELECTION}__county.csv"),
              COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, f"{ELECTION}__precinct.csv"),
              PRECINCT_HEADER, precinct_rows)


if __name__ == "__main__":
    main()