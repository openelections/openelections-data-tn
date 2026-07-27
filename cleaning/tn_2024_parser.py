"""
Parser for the 2024 Tennessee elections, producing county- and precinct-level
OpenElections CSVs for the three 2024 election dates:

  20240305__tn__primary   -- Mar 5 Presidential Preference Primary (presidential
                             preference + the judicial primaries on the March
                             ballot + Republican delegate slates)
  20240801__tn__primary   -- Aug 1 State Primary and General Election: the full
                             Aug 1 ballot (Democratic + Republican primaries,
                             judicial retention, and the 13 judicial "State
                             General" seats decided that day)
  20241105__tn__general   -- Nov 5 General Election (President, U.S. Senate,
                             U.S. House, State Senate, State House)

Source: TN SoS "All by Precinct" spreadsheets (one row per precinct/candidate-
group with up to 10 candidates per row in RNAME*/PARTY*/PVTALLY* columns; a
precinct may span multiple rows -- the CANDGROUP column -- when a race has more
than 10 candidates, e.g. the Republican delegate slates). Grand totals were
cross-checked against the official by-county PDFs.

  https://sos-prod.tnsosgovfiles.com/s3fs-public/document/20240305AllbyPrecinct.xlsx
  https://sos-prod.tnsosgovfiles.com/s3fs-public/document/20240801AllbyPrecinct.xlsx
  https://sos-prod.tnsosgovfiles.com/s3fs-public/document/20241105AllbyPrecinct.xlsx

The source spreadsheet names are listed at https://sos.tn.gov/elections/results#2024
(WebFetch returns 403; use curl with a browser User-Agent).

Conventions (standard OpenElections office names; Title-case counties as
printed in the source; full party names; precinct names as printed):
  "Candidates for President of the United States"  -> office "Presidential Preference", district "NA"
  "President and Vice President of the United States" -> office "President", district "NA",
       candidate shortened from "Electors for <name> for President and ..." to "<name>"
       (matching the repo's 2020 general file)
  "United States Senate"                           -> office "U.S. Senate", district "NA"
  "United States House of Representatives District N" -> office "U.S. House", district "N"
  "Tennessee Senate District N"                    -> office "State Senate", district "N"
  "Tennessee House of Representatives District N"  -> office "State House", district "N"
  "Delegate At-Large" / "Delegate District N"       -> office "Delegate", district "At-Large" / "N"
  "Supreme Court" / "Court of Criminal Appeals - ..." -> office as printed, district "NA"
  "<office> Division/Part ... District N (unexpired term)" -> office = text before "District N",
       district = "N (unexpired term)" (or "N" for full-term races), Division/Part kept in office
  "State Executive Committeeman/woman District N (unexpired term)" -> split likewise

Judicial retention races are non-partisan (no party in the source); the candidate
is kept verbatim as printed ("Retain - <judge>" / "Replace - <judge>") and the
party column is left empty.

County-level totals are derived by summing the precinct votes per
(county, office, district, party, candidate), which matches the by-county PDFs.

Requires `requests` (in the Pipfile) and `openpyxl` (install separately, e.g.
`pip install openpyxl`) to read the .xlsx workbooks.
"""

import csv
import os
import re
from collections import defaultdict

import requests

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("This parser requires openpyxl: pip install openpyxl") from exc

BASE = "https://sos-prod.tnsosgovfiles.com/s3fs-public/document"

ELECTIONS = [
    {"date": "20240305", "type": "primary",
     "url": f"{BASE}/20240305AllbyPrecinct.xlsx"},
    {"date": "20240801", "type": "primary",
     "url": f"{BASE}/20240801AllbyPrecinct.xlsx"},
    {"date": "20241105", "type": "general",
     "url": f"{BASE}/20241105AllbyPrecinct.xlsx"},
]

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2024")

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party", "candidate", "votes"]

USHOUSE_RE = re.compile(r"^United States House of Representatives District (\d+)\s*$")
TNSENATE_RE = re.compile(r"^Tennessee Senate District (\d+)\s*$")
TNHOUSE_RE = re.compile(r"^Tennessee House of Representatives District (\d+)\s*$")
DELEGATE_DIST_RE = re.compile(r"^Delegate District (\d+)\s*$")
DISTRICT_RE = re.compile(r"^(.+?)\s+District\s+(\d+)\s*(\(unexpired term\))?\s*$")
ELECTORS_RE = re.compile(r"^Electors for (.+?) for President and .+ for Vice President\s*$")
WHITESPACE_RE = re.compile(r"\s+")


def clean(s):
    """Strip and collapse internal whitespace (the file_format test rejects
    consecutive whitespace, and some source precinct names contain double
    spaces, e.g. "13-29  Polk Clark")."""
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def norm_office(raw):
    """Map a source OFFICENAME to (office, district). Returns None to skip."""
    s = raw.strip()
    if s == "Candidates for President of the United States":
        return ("Presidential Preference", "NA")
    if s == "President and Vice President of the United States":
        return ("President", "NA")
    if s == "United States Senate":
        return ("U.S. Senate", "NA")
    m = USHOUSE_RE.match(s)
    if m:
        return ("U.S. House", m.group(1))
    m = TNSENATE_RE.match(s)
    if m:
        return ("State Senate", m.group(1))
    m = TNHOUSE_RE.match(s)
    if m:
        return ("State House", m.group(1))
    if s == "Delegate At-Large":
        return ("Delegate", "At-Large")
    m = DELEGATE_DIST_RE.match(s)
    if m:
        return ("Delegate", m.group(1))
    if s == "Supreme Court" or s.startswith("Court of Appeals") \
            or s.startswith("Court of Criminal Appeals"):
        return (s, "NA")
    m = DISTRICT_RE.match(s)
    if m:
        office = m.group(1).strip()
        dist = m.group(2) + (" (unexpired term)" if m.group(3) else "")
        return (office, dist)
    # Fallback: keep the source office name, no district.
    return (s, "NA")


def norm_candidate(raw_office, candidate):
    c = str(candidate).strip()
    if raw_office.strip() == "President and Vice President of the United States":
        m = ELECTORS_RE.match(c)
        if m:
            return m.group(1)
    return c


def fetch_workbook(url):
    resp = requests.get(url)
    resp.raise_for_status()
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(resp.content)
        f.flush()
        return openpyxl.load_workbook(f.name, read_only=True, data_only=True)


def iter_rows(url):
    """Yield (county, precinct, office_raw, party, candidate, votes)."""
    wb = fetch_workbook(url)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ci_county = header.index("COUNTY")
    ci_precinct = header.index("PRECINCT")
    ci_office = header.index("OFFICENAME")
    rname_cols = [header.index(f"RNAME{i}") for i in range(1, 11)
                  if f"RNAME{i}" in header]
    party_cols = [header.index(f"PARTY{i}") for i in range(1, 11)
                  if f"PARTY{i}" in header]
    tally_cols = [header.index(f"PVTALLY{i}") for i in range(1, 11)
                  if f"PVTALLY{i}" in header]
    n = min(len(rname_cols), len(party_cols), len(tally_cols))
    for r in rows:
        county = clean(r[ci_county] or "")
        precinct = clean(r[ci_precinct] or "")
        office_raw = clean(r[ci_office] or "")
        for k in range(n):
            name = r[rname_cols[k]]
            if not name:
                continue
            party = clean(r[party_cols[k]] or "")
            votes = r[tally_cols[k]] or 0
            yield (county, precinct, office_raw, party,
                   clean(norm_candidate(office_raw, name)), int(votes))


def build(election):
    """Build county + precinct rows for one election; return (county_rows, precinct_rows)."""
    precinct_rows = []
    county_sums = defaultdict(int)  # (county, office, district, party, candidate) -> votes
    seen_precinct = set()
    for county, precinct, office_raw, party, candidate, votes in iter_rows(election["url"]):
        nd = norm_office(office_raw)
        if nd is None:
            continue
        office, district = nd
        key = (county, precinct, office, district, party, candidate)
        if key in seen_precinct:
            # Should not happen; a duplicate would mean a parse/source issue.
            raise ValueError(f"duplicate precinct row: {key}")
        seen_precinct.add(key)
        precinct_rows.append([county, precinct, office, district, party, candidate, votes])
        county_sums[(county, office, district, party, candidate)] += votes

    county_rows = [
        [county, office, district, party, candidate, votes]
        for (county, office, district, party, candidate), votes in county_sums.items()
    ]
    return county_rows, precinct_rows, county_sums


def sort_key_county(row):
    return (row[0], row[1], row[2], row[3], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"Wrote {path} ({len(rows)} rows)")


def main():
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)

    for election in ELECTIONS:
        date = election["date"]
        etype = election["type"]
        print(f"--- {date} {etype} ---")
        county_rows, precinct_rows, county_sums = build(election)
        county_rows.sort(key=sort_key_county)
        precinct_rows.sort(key=sort_key_precinct)
        print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows")

        write_csv(os.path.join(out, f"{date}__tn__{etype}__county.csv"),
                  COUNTY_HEADER, county_rows)
        write_csv(os.path.join(out, f"{date}__tn__{etype}__precinct.csv"),
                  PRECINCT_HEADER, precinct_rows)

        # Sanity: precinct sums must equal county totals.
        ps = defaultdict(int)
        for r in precinct_rows:
            ps[(r[0], r[2], r[3], r[4], r[5])] += r[6]
        assert set(ps) == set(county_sums), \
            f"{date}: precinct keys != county keys"
        for k, v in county_sums.items():
            assert ps[k] == v, f"{date}: mismatch {k}: {ps[k]} != {v}"
        print(f"  OK: precinct sums equal county totals "
              f"({len(county_sums)} combinations)")


if __name__ == "__main__":
    main()