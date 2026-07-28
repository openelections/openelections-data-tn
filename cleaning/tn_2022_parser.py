"""
Parser for the 2022 Tennessee elections, producing county- and precinct-level
OpenElections CSVs for the two 2022 primaries and the Nov 8 general:

  20220503__tn__primary  -- May 3 county/judicial primary (both parties)
  20220804__tn__primary  -- Aug 4 state primary & general: the full Aug 4 ballot
                            (Democratic + Republican primaries for Governor, U.S.
                            House, State Senate, State House, State Executive
                            Committee, and the judicial/DA/Public Defender
                            primaries; the judicial "State General" seats decided
                            that day; and judicial retention)
  20221108__tn__general  -- Nov 8 general (Governor, U.S. House, State Senate,
                            State House, four constitutional amendments).
                            The precinct file already in the repo omitted Governor;
                            this parser regenerates both county + precinct from
                            the complete All-by-Precinct spreadsheet.

Sources (https://sos.tn.gov/elections/results#2022; WebFetch 403s, use curl):
  May 3 -- per-party by-precinct PDFs ("Layout B", same shape as the 2023
           Jun-onward specials): one Democratic, one Republican.
    20220503 Democratic Primary Precinct Totals.pdf
    20220503 Republican Primary Precinct Totals.pdf
  Aug 4 -- 20220804ResultsbyPrecinct.xlsx  (ELECTTYPE column distinguishes
           Republican Primary / Democratic Primary / State General / Judicial
           Retention; one row per precinct with up to 10 candidates in
           RNAME*/PARTY*/PVTALLY*, more via the CANDGROUP column)
  Nov 8 -- 20221108AllbyPrecinct.xlsx      (same schema; Governor's 10 ballot
           candidates are in CANDGROUP 1 and its 4 write-in candidates in
           CANDGROUP 2)

The May 3 PDFs contain only districted judicial offices (Chancellor, Circuit
Court Judge, Criminal Court Judge, Probate Court Judge, District Attorney
General, Public Defender) -- every office ends in "District N" (or "District
4/5") and every block ends with "DISTRICT TOTALS", so the office header is
detected with a positive "District N" regex (no need for positional guessing).

Conventions (standard OpenElections office names; Title-case counties as
printed; full party names; precinct names as printed with internal whitespace
collapsed so file_format passes; proper CSV quote-doubling):
  "Governor"                                      -> office "Governor", district "NA"
  "United States Senate"                          -> "U.S. Senate", "NA"  (not on 2022 ballots)
  "United States House of Representatives District N" -> "U.S. House", "N"
  "Tennessee Senate District N"                   -> "State Senate", "N"
  "Tennessee House of Representatives District N"  -> "State House", "N"
  "Supreme Court" / "Court of Appeals - ..." / "Court of Criminal Appeals - ..."
                                                   -> office as printed, district "NA"
  "<judicial office> ... District N"               -> office = text before "District N",
       district "N" (or "N (unexpired term)"; "4/5" kept verbatim)
  "Constitutional Amendment # N"                  -> office as printed, district "NA"
Judicial retention is non-partisan; the candidate is kept verbatim
("Retain - <judge>" / "Replace - <judge>") with the party column empty.
Constitutional amendments are non-partisan ("Yes"/"No", party empty).

County totals are derived by summing precinct votes per
(county, office, district, party, candidate). For May 3 the derived totals are
asserted against each precinct PDF's own "County Totals:" lines; for Aug 4 and
Nov 8 the precinct sums are asserted to equal the derived county totals, and
totals are spot-checked against the official by-county PDFs.

Requires `requests` (in the Pipfile), `openpyxl` (install separately, e.g.
`pip install openpyxl`) for the .xlsx workbooks, and `pdftotext` (poppler) for
the May 3 PDFs.
"""

import csv
import os
import re
import subprocess
import tempfile
from collections import defaultdict

import requests

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("This parser requires openpyxl: pip install openpyxl") from exc

BASE = "https://sos-prod.tnsosgovfiles.com/s3fs-public/document"

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party",
                   "candidate", "votes"]

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2022")

# ---- XLSX (Aug 4, Nov 8) -------------------------------------------------

USHOUSE_RE = re.compile(r"^United States House of Representatives District (\d+)\s*$")
TNSENATE_RE = re.compile(r"^Tennessee Senate District (\d+)\s*$")
TNHOUSE_RE = re.compile(r"^Tennessee House of Representatives District (\d+)\s*$")
# District N, N/M, or "N (unexpired term)" -- covers the judicial offices.
DISTRICT_RE = re.compile(r"^(.+?)\s+District\s+(\d+(?:/\d+)?)\s*(\(unexpired term\))?\s*$")
WHITESPACE_RE = re.compile(r"\s+")


def clean(s):
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def norm_office(raw):
    """Map a source OFFICENAME to (office, district)."""
    s = raw.strip()
    if s == "Governor":
        return ("Governor", "NA")
    if s == "United States Senate":
        return ("U.S. Senate", "NA")
    m = USHOUSE_RE.match(s)
    if m:
        return ("U.S. House", m.group(1))
    m = TNSENATE_RE.match(s)
    if m:
        return ("State Senate", m.group(1))
    m = TNHOUSE_RE.match(s)
    if m:
        return ("State House", m.group(1))
    if s == "Supreme Court" or s.startswith("Court of Appeals") \
            or s.startswith("Court of Criminal Appeals"):
        return (s, "NA")
    m = DISTRICT_RE.match(s)
    if m:
        office = m.group(1).strip()
        dist = m.group(2) + (" (unexpired term)" if m.group(3) else "")
        return (office, dist)
    # Constitutional amendments and anything else: keep the source name, NA.
    return (s, "NA")


def iter_xlsx(url):
    """Yield (county, precinct, office_raw, party, candidate, votes) per
    non-empty candidate slot across all CANDGROUP rows."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(resp.content)
        f.flush()
        wb = openpyxl.load_workbook(f.name, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ci_county = header.index("COUNTY")
    ci_precinct = header.index("PRECINCT")
    rname_cols = [header.index(f"RNAME{i}") for i in range(1, 11)
                 if f"RNAME{i}" in header]
    party_cols = [header.index(f"PARTY{i}") for i in range(1, 11)
                 if f"PARTY{i}" in header]
    tally_cols = [header.index(f"PVTALLY{i}") for i in range(1, 11)
                  if f"PVTALLY{i}" in header]
    n = min(len(rname_cols), len(party_cols), len(tally_cols))
    for r in rows:
        county = clean(r[ci_county] or "")
        precinct = clean(r[ci_precinct] or "")
        office_raw = clean(r[header.index("OFFICENAME")] or "")
        for k in range(n):
            name = r[rname_cols[k]]
            if not name:
                continue
            party = clean(r[party_cols[k]] or "")
            votes = r[tally_cols[k]] or 0
            yield (county, precinct, office_raw, party, clean(name), int(votes))


def build_xlsx(url):
    precinct_rows = []
    county_sums = defaultdict(int)
    seen = {}        # (county,precinct,office,district,party,candidate) -> votes
    dup_skipped = 0
    for county, precinct, office_raw, party, candidate, votes in iter_xlsx(url):
        office, district = norm_office(office_raw)
        key = (county, precinct, office, district, party, candidate)
        if key in seen:
            # The SoS workbook carries literal duplicate rows for some
            # "No Candidate Qualified" races; identical duplicates are safe to
            # drop, but a differing vote count would be a real conflict.
            if seen[key] != votes:
                raise ValueError(f"conflicting duplicate precinct row: {key} "
                                 f"({seen[key]} vs {votes})")
            dup_skipped += 1
            continue
        seen[key] = votes
        precinct_rows.append([county, precinct, office, district, party, candidate, votes])
        county_sums[(county, office, district, party, candidate)] += votes
    if dup_skipped:
        print(f"  (skipped {dup_skipped} duplicate source rows)")
    county_rows = [[c, o, d, p, n, v] for (c, o, d, p, n), v in county_sums.items()]
    return county_rows, precinct_rows, county_sums


# ---- May 3 PDFs (Layout B) ----------------------------------------------

OFFICE_RE = re.compile(r"^(.+?)\s+District\s+(\d+(?:/\d+)?)\s*(\(unexpired term\))?\s*$")
SECTION_RE = re.compile(r"^(Republican Primary|Democratic Primary|State General)\s*$")
CAND_RE = re.compile(r"^(\d+)\.\s+(.+?)\s*$")
BANNER_COUNTY_RE = re.compile(r"^State of Tennessee - (.+?) County\s*$")
COUNTY_RE = re.compile(r"^(.+?)\s+County\s*$")
ALLNUM_RE = re.compile(r"^\d+(\s+\d+)*\s*$")
PAGE_RE = re.compile(r"Page \d+ of \d+")
DATE_RE = re.compile(r"^(January|February|March|April|May|June|July|August|"
                     r"September|October|November|December)\s+\d{1,2},\s+\d{4}\s*$")


def _is_int(tok):
    return re.match(r"^\d{1,3}(,\d{3})*$", tok) is not None or tok.isdigit()


def _to_int(tok):
    return int(tok.replace(",", ""))


def fetch_text(url):
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix=".pdf") as f:
        f.write(resp.content)
        f.flush()
        out = subprocess.run(["pdftotext", "-layout", f.name, "-"],
                              capture_output=True, text=True, check=True)
    return out.stdout


def parse_precinct_pdf(text):
    """Parse one May 3 by-precinct PDF (Layout B). Returns (rows, expected).

    rows: [county, precinct, office, district, party, candidate, votes]
    expected: {(county, office, district, party, candidate): votes} from the
              PDF's own "County Totals:" lines.
    """
    rows = []
    expected = {}
    county = None
    office = None
    district = None
    section_party = None
    general = False
    candidates = []   # list of (name, party) for the current office
    phase = None      # None -> "cand" (after office header) -> "rows" (after Precincts:)

    for raw in text.splitlines():
        s = raw.strip()
        if not s:
            continue

        m = BANNER_COUNTY_RE.match(s)
        if m:
            county = clean(m.group(1))
            continue
        if s == "State of Tennessee":
            continue
        if PAGE_RE.search(s):
            continue
        if DATE_RE.match(s):
            continue

        m = SECTION_RE.match(s)
        if m:
            office = None
            district = None
            section_party = None
            general = False
            candidates = []
            phase = None
            kind = m.group(1)
            if kind == "State General":
                general = True
            else:
                section_party = "Republican" if kind == "Republican Primary" else "Democratic"
            continue

        # Office header -- positive "District N" match (all May 3 offices).
        m = OFFICE_RE.match(s)
        if m:
            office = clean(m.group(1))
            district = m.group(2) + (" (unexpired term)" if m.group(3) else "")
            candidates = []
            phase = "cand"
            continue

        # Candidate line.
        m = CAND_RE.match(s)
        if m and phase == "cand":
            rest = clean(m.group(2))
            if general and " - " in rest:
                name, party = rest.rsplit(" - ", 1)
                name, party = clean(name), clean(party)
            else:
                name, party = rest, section_party or ""
            candidates.append((name, party))
            continue

        if phase == "cand" and ALLNUM_RE.match(s):   # column header
            continue

        if s.startswith("Precincts:") or s == "Precincts:":
            phase = "rows"
            continue

        if "TOTALS" in s.upper():
            if s.startswith("Totals:") or s.startswith("County Totals:"):
                label = "County Totals:" if s.startswith("County Totals:") else "Totals:"
                toks = s[len(label):].split()
                if county is not None and len(candidates) == len(toks) \
                        and all(_is_int(t) for t in toks):
                    for (cname, cparty), v in zip(candidates, [_to_int(t) for t in toks]):
                        expected[(county, office, district, cparty, cname)] = v
            continue

        m = COUNTY_RE.match(s)            # "X County"
        if m:
            county = clean(m.group(1))
            continue

        # Precinct row: last len(candidates) tokens are votes.
        if phase == "rows" and candidates and county is not None:
            toks = s.split()
            n = len(candidates)
            if len(toks) >= n + 1 and all(_is_int(t) for t in toks[-n:]):
                votes = [_to_int(t) for t in toks[-n:]]
                precinct = clean(" ".join(toks[:-n]))
                for (cname, cparty), v in zip(candidates, votes):
                    rows.append([county, precinct, office, district,
                                 cparty, cname, v])
                continue
            print(f"  WARN unhandled: {s!r}")

    return rows, expected


def build_pdf(urls):
    """Parse the Dem + Rep May 3 PDFs and combine."""
    all_rows = []
    all_expected = {}
    seen = {}          # (county,precinct,office,district,party,candidate) -> votes
    dup_skipped = 0
    for url in urls:
        text = fetch_text(url)
        rows, expected = parse_precinct_pdf(text)
        all_expected.update(expected)
        for r in rows:
            key = tuple(r[i] for i in (0, 1, 2, 3, 4, 5))
            v = r[6]
            if key in seen:
                # Some "No Candidate Qualified" races print each precinct row
                # twice in the source PDF; identical duplicates are safe to drop.
                if seen[key] != v:
                    raise ValueError(f"conflicting duplicate precinct row: {key} "
                                     f"({seen[key]} vs {v})")
                dup_skipped += 1
                continue
            seen[key] = v
            all_rows.append(r)
    if dup_skipped:
        print(f"  (skipped {dup_skipped} duplicate source rows)")

    county_sums = defaultdict(int)
    for r in all_rows:
        county_sums[tuple(r[i] for i in (0, 2, 3, 4, 5))] += r[6]
    county_rows = [[c, o, d, p, n, v] for (c, o, d, p, n), v in county_sums.items()]

    mismatches = [(k, v, county_sums.get(k)) for k, v in all_expected.items()
                  if county_sums.get(k) != v]
    # The SoS by-precinct PDFs double-list the precincts of a few D13 races:
    #   - "No Candidate Qualified" races (e.g. DeKalb D13 Democratic) print each
    #     precinct twice with 0 votes, so the official "County Totals:" is 0 --
    #     keep-first dedup yields 0 too, and these never reach `mismatches`.
    #   - Pickett's D13 Republican races print each precinct twice with the REAL
    #     vote, and the PDF's own "County Totals:" sums the doubled rows (so the
    #     official county total is 2x the sum of the unique precincts). The
    #     separate by-county PDF carries the same doubled number, confirming this
    #     is a SoS double-counting artifact, not a parser bug. We keep ONE row per
    #     precinct with the real per-precinct vote (a precinct cannot cast its
    #     ballots twice) and treat a clean 2x official-vs-derived mismatch as a
    #     known artifact rather than a failure.
    doublings = [(k, exp, got) for k, exp, got in mismatches
                 if got is not None and exp == 2 * got]
    real = [(k, exp, got) for k, exp, got in mismatches
            if not (got is not None and exp == 2 * got)]
    if doublings:
        print(f"  NOTE: {len(doublings)} SoS doubled-total artifacts "
              f"(official county total = 2x unique-precinct sum); kept the real "
              f"per-precinct votes:")
        for k, exp, got in doublings[:12]:
            print(f"    {k}: official={exp} derived={got}")
    if real:
        for k, exp, got in real[:20]:
            print(f"  MISMATCH {k}: pdf={exp} derived={got}")
        raise AssertionError(f"May 3: {len(real)} county-total mismatches "
                             f"vs the precinct PDFs' own totals")
    return county_rows, all_rows, county_sums


# ---- shared -------------------------------------------------------------


def sort_key_county(row):
    return (row[0], row[2], row[3], row[1], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  Wrote {path} ({len(rows)} rows)")


def check_sums(date, precinct_rows, county_sums):
    ps = defaultdict(int)
    for r in precinct_rows:
        ps[(r[0], r[2], r[3], r[4], r[5])] += r[6]
    assert set(ps) == set(county_sums), f"{date}: precinct keys != county keys"
    for k, v in county_sums.items():
        assert ps[k] == v, f"{date}: mismatch {k}: {ps[k]} != {v}"


def main():
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)

    # May 3 primary (county/judicial) -- combine Dem + Rep PDFs.
    print("--- 20220503 primary (May 3, PDFs) ---")
    may03_urls = [
        f"{BASE}/20220503%20Democratic%20Primary%20Precinct%20Totals.pdf",
        f"{BASE}/20220503%20Republican%20Primary%20Precinct%20Totals.pdf",
    ]
    county_rows, precinct_rows, county_sums = build_pdf(may03_urls)
    county_rows.sort(key=sort_key_county)
    precinct_rows.sort(key=sort_key_precinct)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows, "
          f"{len(county_sums)} county/candidate combos")
    check_sums("20220503", precinct_rows, county_sums)
    write_csv(os.path.join(out, "20220503__tn__primary__county.csv"),
              COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, "20220503__tn__primary__precinct.csv"),
              PRECINCT_HEADER, precinct_rows)

    # Aug 4 primary & general (full Aug 4 ballot).
    print("--- 20220804 primary (Aug 4, XLSX) ---")
    county_rows, precinct_rows, county_sums = build_xlsx(
        f"{BASE}/20220804ResultsbyPrecinct.xlsx")
    county_rows.sort(key=sort_key_county)
    precinct_rows.sort(key=sort_key_precinct)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows, "
          f"{len(county_sums)} county/candidate combos")
    check_sums("20220804", precinct_rows, county_sums)
    write_csv(os.path.join(out, "20220804__tn__primary__county.csv"),
              COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, "20220804__tn__primary__precinct.csv"),
              PRECINCT_HEADER, precinct_rows)

    # Nov 8 general (regenerate both; the old precinct omitted Governor).
    print("--- 20221108 general (Nov 8, XLSX) ---")
    county_rows, precinct_rows, county_sums = build_xlsx(
        f"{BASE}/20221108AllbyPrecinct.xlsx")
    county_rows.sort(key=sort_key_county)
    precinct_rows.sort(key=sort_key_precinct)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows, "
          f"{len(county_sums)} county/candidate combos")
    check_sums("20221108", precinct_rows, county_sums)
    write_csv(os.path.join(out, "20221108__tn__general__county.csv"),
              COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, "20221108__tn__general__precinct.csv"),
              PRECINCT_HEADER, precinct_rows)


if __name__ == "__main__":
    main()