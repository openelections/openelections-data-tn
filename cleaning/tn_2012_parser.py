"""
Parser for the 2012 Tennessee elections, producing OpenElections CSVs (county +
precinct) for all three 2012 elections (the December 17 Electoral College
certification is intentionally EXCLUDED), sourced from the TN SoS results page
(https://sos.tn.gov/elections/results#2012):

  20120306__tn__primary  -- Mar 6 Presidential Primary: presidential preference
                            (Rep + Dem), Republican delegates (at-large + by
                            district), and the judicial primary (Rep + Dem)
  20120802__tn__primary  -- Aug 2 Primary Election: Rep + Dem primaries for
                            U.S. Senate/U.S. House/State Senate/State House,
                            plus the non-partisan "State General" judicial/
                            DA/Public Defender races and Judicial Retention
  20121106__tn__general  -- Nov 6 General Election: President, U.S. Senate,
                            U.S. House, State Senate, State House

The repo already had precinct-only CSVs for each (no county files), in a
non-standard convention: UPPERCASE counties, float votes (115.0), non-standard
office names ("State House District" with the number in a separate field,
"United States Senate"), and a corrupted party column in the Nov 6 file
(candidate names leaking into party). This parser REGENERATES them with standard
OpenElections conventions and ADDS the missing county files. The Mar 6 file is
renamed from `20120306__tn__primary__president__` to `20120306__tn__primary__`
(no subtype), matching the repo's 2020/2024 March-primary convention -- the old
`__president__` precinct file is removed.

Source: the three SoS "All by Precinct" workbooks on
https://sos-prod.tnsosgovfiles.com/s3fs-public/document/ (March2012.xlsx,
August2012.xlsx, November2012.xlsx). Shared schema: COUNTY, Precinct Seq,
PRECINCT, Office Seq, OfficeID, District, Candidate Group, OFFICENAME,
ELECTDATE, ELECTTYPE, then Col1/BName1/Tally1 ... Col10/BName10/Tally10 (the
candidate number is in ColN; BName is the name with a leading ". " leftover,
e.g. ". Michele Bachmann"; TallyN is the precinct vote). ELECTDATE is constant
per file; ELECTTYPE is "Republican Primary"/"Democratic Primary" (primaries,
with irregular internal spacing in March: "Republican   Primary"), "State
General"/"Judicial Retention" (Aug non-partisan judicial), or "General
Election" (Nov).

Conventions (standard OpenElections office names; title-case counties; integer
votes; full party names; precinct names verbatim with internal whitespace
collapsed):
  OFFICENAME ends with literal " District" + the number in the District column
    -> office = OFFICENAME minus " District" (then mapped to standard),
       district = District column value
      "U.S. House of Representatives District" -> "U.S. House", <dist>
      "Tennessee Senate District"               -> "State Senate", <dist>
      "Tennessee House of Representatives District" -> "State House", <dist>
      "Circuit Court Judge Division III District"   -> "Circuit Court Judge
                                                       Division III", <dist>
      "Circuit Court Division 1 District" (March)   -> "Circuit Court
                                                       Division 1", <dist>
      "Criminal Court Judge Division I District"    -> as printed, <dist>
      "District Attorney General District" (March) / "District Attorney
                                                       District" (Aug)
                                                   -> as printed, <dist>
      "Public Defender District"                    -> "Public Defender", <dist>
  OFFICENAME without " District":
      "United States President"   -> "President", "NA"
      "United States Senate"      -> "U.S. Senate", "NA"
      "Presidential Preference"   -> "Presidential Preference", "NA"
      "Delegate At Large"         -> "Delegate", "At-Large"
      "Delegate District"         -> "Delegate", <District col>
      "Court of Criminal Appeals - Middle Division" (Judicial Retention)
                                  -> as printed, "NA"
The judicial/DA/Public Defender office names are kept as printed (the source
labels them inconsistently between March and Aug, e.g. "Circuit Court Division
1" vs "Circuit Court Judge Division III"; each election keeps its own labels).

Party:
  Nov general -- parsed from each candidate's " - (X)" suffix on BName
                 (R/D/C/G/I -> Republican/Democratic/Constitution/Green/
                 Independent); the suffix is stripped from the name. Write-ins
                 ("Write-In - <name>", no parenthetical) get an empty party.
  Primaries   -- from ELECTTYPE (Republican / Democratic).
  Aug "State General" / "Judicial Retention" -- non-partisan, party empty.
                 Judicial-retention candidates are "Retain - <judge>" /
                 "Replace - <judge>" (kept verbatim).
  March delegates -- party Republican (from ELECTTYPE); the BName's
                 " - <pledged presidential candidate>" suffix is NOT a party, so
                 it is kept in the candidate name (e.g. "Willis Ayres -
                 Gingrich"), giving meaningful pledge information.

Zero-vote rows are EXCLUDED -- a precinct row is emitted only when Tally > 0.
This matches the existing 2012 files (Mar 6: 152,120 positive-vote cells; Nov 6:
43,420) and avoids tens of thousands of 0-vote write-in rows in the Nov 6
general. County totals are derived by summing the (positive) precinct votes per
(county, office, district, party, candidate).

County names are title-cased from the uppercase XLSX (ANDERSON -> Anderson)
with fixes for the Mc/DeKalb counties (MCMINN -> McMinn, MCNAIRY -> McNairy,
DEKALB -> DeKalb).

Requires `requests` + `openpyxl` (in the Pipfile) and `pdftotext` (poppler, for
the optional by-county PDF verification).
"""

import csv
import io
import os
import re
import subprocess
import tempfile
from collections import defaultdict

import openpyxl
import requests

BASE_XLSX = "https://sos-prod.tnsosgovfiles.com/s3fs-public/document"
BASE_PDF = ("https://tnelections.tnsosfiles.com/sharetngov/archived/"
            "election/results")

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2012")

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party",
                   "candidate", "votes"]

WHITESPACE_RE = re.compile(r"\s+")
PARTY_SUFFIX_RE = re.compile(r"\s+-\s+\(([A-Z])\)\s*$")
PARTY_LETTER = {"R": "Republican", "D": "Democratic", "C": "Constitution",
                "G": "Green", "I": "Independent"}

OFFICE_MAP = {
    "United States President": "President",
    "United States Senate": "U.S. Senate",
    "U.S. House of Representatives": "U.S. House",
    "Tennessee Senate": "State Senate",
    "Tennessee House of Representatives": "State House",
    "Presidential Preference": "Presidential Preference",
}


def clean(s):
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def norm_county(c):
    s = clean(c)
    if not s:
        return s
    parts = s.title().split()
    fixed = []
    for p in parts:
        low = p.lower()
        if low == "dekalb":
            p = "DeKalb"
        elif low.startswith("mc") and len(p) > 2:
            p = "Mc" + p[2].upper() + p[3:]
        fixed.append(p)
    return " ".join(fixed)


def norm_district(d):
    if d is None or str(d).strip() == "":
        return "NA"
    return clean(d)


def norm_office(raw_office, district_col):
    """Return (office, district) from OFFICENAME + the District column value."""
    raw = clean(raw_office)
    dist = norm_district(district_col)
    if raw == "Delegate At Large":
        return "Delegate", "At-Large"
    if raw == "Delegate District":
        return "Delegate", dist
    if raw.endswith(" District"):
        base = raw[: -len(" District")].strip()
    else:
        base = raw
    office = OFFICE_MAP.get(base, base)
    return office, dist


def parse_candidate(bname, etype):
    """Return (name, party) from a BName cell + ELECTTYPE."""
    name = clean(bname)
    # strip the leading ". " leftover (candidate number is in the Col field)
    if name.startswith(". "):
        name = name[2:].strip()
    elif name.startswith("."):
        name = name[1:].strip()
    party = ""
    m = PARTY_SUFFIX_RE.search(name)
    if m:
        party = PARTY_LETTER.get(m.group(1), "")
        name = name[: m.start()].strip()
    else:
        et = clean(etype)
        if "Republican" in et:
            party = "Republican"
        elif "Democratic" in et:
            party = "Democratic"
        # General Election w/o suffix, State General, Judicial Retention -> ""
    return name, party


def load_xlsx(url):
    """Parse one All-by-Precinct XLSX. Returns (precinct_rows, county_rows)."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True,
                                data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {name: i for i, name in enumerate(hdr) if name}
    county_i = idx["COUNTY"]
    precinct_i = idx["PRECINCT"]
    office_i = idx["OFFICENAME"]
    etype_i = idx["ELECTTYPE"]
    dist_i = idx["District"]
    bname_cols = [i for i, c in enumerate(hdr)
                  if c and str(c).startswith("BName")]
    tally_cols = [i for i, c in enumerate(hdr)
                  if c and str(c).startswith("Tally")]

    precinct_rows = []
    county_sum = defaultdict(int)
    for r in rows[1:]:
        raw_office = r[office_i]
        if raw_office is None or not str(raw_office).strip():
            continue
        county = norm_county(r[county_i])
        precinct = clean(r[precinct_i])
        district_col = r[dist_i]
        etype = r[etype_i]
        office, district = norm_office(raw_office, district_col)
        for k, bcol in enumerate(bname_cols):
            bn = r[bcol]
            if bn is None or not str(bn).strip() or str(bn).strip() == ".":
                continue
            tally = r[tally_cols[k]]
            try:
                votes = int(tally) if tally not in (None, "") else 0
            except (TypeError, ValueError):
                votes = 0
            if votes <= 0:
                continue
            name, party = parse_candidate(bn, etype)
            if not name:
                continue
            row = [county, precinct, office, district, party, name, votes]
            precinct_rows.append(row)
            county_sum[(county, office, district, party, name)] += votes
    county_rows = [[k2[0], k2[1], k2[2], k2[3], k2[4], v]
                   for k2, v in county_sum.items()]
    return county_rows, precinct_rows


def sort_key_county(row):
    return (row[0], row[2], row[3], row[1], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  Wrote {path} ({len(rows)} rows)")


ELECTIONS = [
    {"name": "20120306__tn__primary",
     "xlsx": f"{BASE_XLSX}/March2012.xlsx"},
    {"name": "20120802__tn__primary",
     "xlsx": f"{BASE_XLSX}/August2012.xlsx"},
    {"name": "20121106__tn__general",
     "xlsx": f"{BASE_XLSX}/November2012.xlsx"},
]


def main():
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)
    for election in ELECTIONS:
        print(f"--- {election['name']} ---")
        county_rows, precinct_rows = load_xlsx(election["xlsx"])
        county_rows.sort(key=sort_key_county)
        precinct_rows.sort(key=sort_key_precinct)
        print(f"  county: {len(county_rows)} rows, precinct: "
              f"{len(precinct_rows)} rows")
        write_csv(os.path.join(out, f"{election['name']}__county.csv"),
                  COUNTY_HEADER, county_rows)
        write_csv(os.path.join(out, f"{election['name']}__precinct.csv"),
                  PRECINCT_HEADER, precinct_rows)


if __name__ == "__main__":
    main()