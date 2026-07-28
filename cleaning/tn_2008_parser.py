"""
Parser for the 2008 Tennessee elections, producing OpenElections CSVs (county
+ precinct) for all three 2008 elections, sourced from the TN SoS results page
(https://sos.tn.gov/elections/results#2008):

  20080205__tn__primary  -- Feb 5 Presidential Preference Primary: Democratic +
        Republican presidential preference (the "All by Precinct" workbook
        contains ONLY presidential preference; the Republican delegate slates
        are published as separate PDFs and are NOT in the workbook, so they are
        excluded). The old repo file was named "...__primary__president__"; the
        "__president" subtype is DROPPED to match the 2020/2024 (and 2012 March)
        March-primary convention (bare "__primary").
  20080807__tn__primary  -- Aug 7 primary: Republican + Democratic primaries
        for U.S. Senate / U.S. House / State Senate / State House. (The Aug
        ballot also had judicial retention + a "State General" judicial contest,
        but those are NOT in the "All by Precinct" workbook -- only the partisan
        primaries are -- so they are excluded, matching the old repo file's
        scope.)
  20081104__tn__general  -- Nov 4 general: President, U.S. Senate, U.S. House,
        State Senate, State House. (No constitutional amendments on the 2008
        general ballot.)

The repo already had precinct-only CSVs for all three (no county files) in a
non-standard convention: UPPERCASE counties, float votes (290.0), single-letter
party (R/D) or candidate-name-in-party, and verbose office names ("United States
Senate", "U.S. House of Representatives District", "State House District",
 "President"). This parser REGENERATES them with standard OpenElections
conventions and ADDS the missing county files.

Source workbooks (the SoS "All by Precinct" spreadsheets on
https://sos-prod.tnsosgovfiles.com/s3fs-public/document/): February2008.xlsx,
August2008.xlsx, November2008.xlsx (all .xlsx, read with openpyxl). Shared
schema: COUNTY, Prct Seq, PRECINCT, Office Seq, OfficeID, District, Candidate
Group, OFFICENAME, ELECTDATE, ELECTTYPE, then COL1/BNAME1/TALLY1 ...
COL10/BNAME10/TALLY10 (the candidate number is in COLN; BNAME is the name;
TALLYN is the precinct vote). ELECTTYPE is "Republican Primary"/"Democratic
Primary" (Feb + Aug) or "General Election" (Nov). The District column holds the
string "None" when there is no district. Candidate Group is "1" for every 2008
race (no race has >10 candidates), so there is no overflow handling -- every
non-empty BNAME cell is emitted, capturing each candidate exactly once.

Conventions (standard OpenElections office names; title-case counties; integer
votes; full party names; bare legislative district number; precinct names
verbatim with internal whitespace collapsed):
  office  OFFICENAME ends with " District" + the number in the District column
            -> office = OFFICENAME minus " District" (then mapped to standard),
               district = the District column value:
            "U.S. House of Representatives District" -> "U.S. House", <d>
            "Tennessee Senate District"             -> "State Senate", <d>
            "Tennessee House of Representatives District" -> "State House", <d>
          OFFICENAME without " District":
            "Presidential Preference"  -> "Presidential Preference", "NA"
            "United States President"   -> "President", "NA"
            "United States Senate"      -> "U.S. Senate", "NA"
  party   Primaries (Feb + Aug) -> from ELECTTYPE (Republican / Democratic).
          The Nov general takes party from each candidate's " - (X)" BNAME
          suffix -- WITH parens, unlike 2010's no-parens " - R" (only D/R/I
          appear in the 2008 source; Cynthia McKinney, the Green nominee
          nationally, is listed as "- (I)" in Tennessee). The suffix is stripped
          from the name and mapped to the full party name.
  write-ins  "Write-in - <name>" / "Write-In - <name>" (the source is
            inconsistent on casing) -> normalized to "Write-In - <name>". A
            primary write-in takes the primary's party (Republican/Democratic
            from ELECTTYPE); a Nov general write-in has an empty party (none of
            the 2008 write-ins carry a party suffix).
  "Uncommitted" (a Feb primary ballot option) is kept verbatim, party from
            ELECTTYPE.

Zero-vote rows ARE included (the source lists every candidate per precinct,
including "Uncommitted" and write-ins at 0), matching the dominant regenerated
convention (2010/2014/2016/2018/2024 all include 0-vote rows; only 2012 and 2020
excluded them) and the 2010 sibling (same source format). County totals are the
sum of the precinct rows per (county, office, district, party, candidate).

County names are title-cased from the uppercase source (ANDERSON -> Anderson)
with fixes for the Mc/DeKalb counties (MCMINN -> McMinn, MCNAIRY -> McNairy,
DEKALB -> DeKalb).

Requires `requests` + `openpyxl` (in the Pipfile).
"""

import csv
import io
import os
import re
from collections import defaultdict

import openpyxl
import requests

BASE_XLSX = "https://sos-prod.tnsosgovfiles.com/s3fs-public/document"
FEB_URL = f"{BASE_XLSX}/February2008.xlsx"
AUG_URL = f"{BASE_XLSX}/August2008.xlsx"
NOV_URL = f"{BASE_XLSX}/November2008.xlsx"

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2008")

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party",
                   "candidate", "votes"]

WHITESPACE_RE = re.compile(r"\s+")
# Nov general party suffix, WITH parens ("Barack Obama - (D)"); only treated as
# a party when the letter is a known party (R/D/C/G/I/L) -- see parse_candidate.
PARENS_SUFFIX_RE = re.compile(r"^(.+?)\s+-\s+\(([A-Z])\)\s*$")
PARTY_LETTER = {"R": "Republican", "D": "Democratic", "C": "Constitution",
                "G": "Green", "I": "Independent", "L": "Libertarian"}

OFFICE_MAP = {
    "Presidential Preference": "Presidential Preference",
    "United States President": "President",
    "United States Senate": "U.S. Senate",
    "U.S. House of Representatives": "U.S. House",
    "Tennessee Senate": "State Senate",
    "Tennessee House of Representatives": "State House",
}


def clean(s):
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def norm_county(c):
    s = clean(c)
    if not s:
        return s
    parts = s.title().split()
    fixed = []
    for p in parts:
        low = p.lower()
        if low == "dekalb":
            p = "DeKalb"
        elif low.startswith("mc") and len(p) > 2:
            p = "Mc" + p[2].upper() + p[3:]
        fixed.append(p)
    return " ".join(fixed)


def norm_district(d):
    if d is None:
        return "NA"
    s = clean(d)
    if s == "" or s.lower() == "none":
        return "NA"
    # numeric -> bare number
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


def primary_party(etype):
    et = clean(etype)
    if "Republican" in et:
        return "Republican"
    if "Democratic" in et:
        return "Democratic"
    return ""


def norm_office(raw_office, district_col):
    """Return (office, district) from OFFICENAME + the District column."""
    raw = clean(raw_office)
    if raw.endswith(" District"):
        base = raw[: -len(" District")].strip()
        return OFFICE_MAP.get(base, base), norm_district(district_col)
    return OFFICE_MAP.get(raw, raw), "NA"


def parse_candidate(bname, etype):
    """Return (name, party) from a BNAME cell + ELECTTYPE.

    Write-ins are normalized to "Write-In - <rest>" (the source casing is
    inconsistent: "Write-in"/"Write-In"); a primary write-in takes the primary's
    party, a Nov general write-in is empty. A Nov general candidate carries a
    " - (X)" parens party suffix which is stripped and mapped to the full party
    name. Bare names take party from the ELECTTYPE (primary) or "" (general)."""
    name = clean(bname)
    # strip a leading ". " leftover if present (2008 has none, but be safe)
    if name.startswith(". "):
        name = name[2:].strip()
    elif name.startswith("."):
        name = name[1:].strip()
    if not name:
        return "", ""
    if name.lower().startswith("write-in"):
        rest = name[len("write-in"):].lstrip()
        if rest[:1] == "-":
            rest = rest[1:].lstrip()
        party = primary_party(etype)
        m = PARENS_SUFFIX_RE.match(rest)
        if m and m.group(2) in PARTY_LETTER:
            rest = clean(m.group(1))
            party = PARTY_LETTER[m.group(2)]
        return f"Write-In - {rest}".strip(), party
    m = PARENS_SUFFIX_RE.match(name)
    if m and m.group(2) in PARTY_LETTER:
        return clean(m.group(1)), PARTY_LETTER[m.group(2)]
    return name, primary_party(etype)


def _column_index(header, *candidates):
    for c in candidates:
        if c in header:
            return header.index(c)
    raise KeyError(f"none of {candidates} found in header")


def _iter_workbook(url):
    """Yield (county, precinct, office_raw, district_col, etype, [bname cells],
    [tally cells]) per source row."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True,
                                data_only=True)
    ws = wb[wb.sheetnames[0]]
    allrows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [clean(c) if c is not None else "" for c in allrows[0]]
    county_i = _column_index(header, "COUNTY")
    precinct_i = _column_index(header, "PRECINCT")
    office_i = _column_index(header, "OFFICENAME")
    etype_i = _column_index(header, "ELECTTYPE")
    dist_i = _column_index(header, "District")
    # 2008 workbooks uppercase the column names (BNAME1/TALLY1) unlike the
    # 2010/2012 mixed-case (BName1/Tally1), so match case-insensitively.
    bname_cols = [i for i, c in enumerate(header) if c.upper().startswith("BNAME")]
    tally_cols = [i for i, c in enumerate(header) if c.upper().startswith("TALLY")]
    for r in allrows[1:]:
        yield (r[county_i], r[precinct_i], r[office_i], r[dist_i],
               r[etype_i],
               [r[c] for c in bname_cols],
               [r[c] for c in tally_cols])


def to_int(v):
    if v is None:
        return 0
    if isinstance(v, bool):
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s == "":
        return 0
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return 0


def load_workbook(url):
    """Parse one All-by-Precinct workbook. Returns (precinct_rows, county_rows).
    Every non-empty BNAME cell is emitted. Zero-vote rows are included."""
    precinct_rows = []
    county_sum = defaultdict(int)
    for (county, precinct, office_raw, district_col, etype,
         bnames, tallies) in _iter_workbook(url):
        if office_raw is None or not str(office_raw).strip():
            continue
        county = norm_county(county)
        precinct = clean(precinct)
        office, district = norm_office(office_raw, district_col)
        for k, bn in enumerate(bnames):
            if bn is None or not str(bn).strip() or str(bn).strip() == ".":
                continue
            name, party = parse_candidate(bn, etype)
            if not name:
                continue
            votes = to_int(tallies[k])
            precinct_rows.append([county, precinct, office, district, party,
                                  name, votes])
            county_sum[(county, office, district, party, name)] += votes
    county_rows = [[c, o, d, p, n, v]
                   for (c, o, d, p, n), v in county_sum.items()]
    return precinct_rows, county_rows


def sort_key_county(row):
    return (row[0], row[1], row[2], row[3], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  Wrote {path} ({len(rows)} rows)")


def build(name, url):
    print(f"--- {name} ---")
    precinct_rows, county_rows = load_workbook(url)
    precinct_rows.sort(key=sort_key_precinct)
    county_rows.sort(key=sort_key_county)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows")
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)
    write_csv(os.path.join(out, f"{name}__county.csv"), COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, f"{name}__precinct.csv"), PRECINCT_HEADER,
              precinct_rows)


def main():
    build("20080205__tn__primary", FEB_URL)
    build("20080807__tn__primary", AUG_URL)
    build("20081104__tn__general", NOV_URL)


if __name__ == "__main__":
    main()