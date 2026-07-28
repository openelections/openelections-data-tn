"""
Parser for the 2020 Tennessee elections, producing OpenElections CSVs for the
two 2020 primaries (county + precinct) and the Nov 3 general (county only; the
precinct file already exists in the repo and is left untouched):

  20200303__tn__primary  -- Mar 3 Presidential Preference Primary: the full
                            Mar 3 ballot (Democratic + Republican presidential
                            preference, the Republican delegate slates, and the
                            judicial primaries on the March ballot)
  20200806__tn__primary  -- Aug 6 State Primary & General: the full Aug 6
                            ballot (Democratic + Republican primaries for U.S.
                            Senate, U.S. House, State Senate, State House, and
                            State Executive Committeeman/woman; the judicial
                            "State General" seats decided that day; and judicial
                            retention)
  20201103__tn__general  -- Nov 3 general, COUNTY ONLY. The precinct file
                            already in the repo (20201103__tn__general__precinct.csv)
                            is left untouched; the county file is derived by
                            summing that precinct file so it matches the
                            existing precinct's conventions exactly.

Sources (https://sos.tn.gov/elections/results#2020; WebFetch 403s, use curl):
  Mar 3 -- March2020Results.xlsx
  Aug 6 -- Aug2020PrecinctDetail.xlsx
  Nov 3 -- Nov2020PrecinctDetail.xlsx  (used only as a reference for the
            by-county spot-check; the county CSV is derived from the existing
            precinct file per the user's instruction)
All on the sos-tn-gov-files bucket (https://sos-tn-gov-files.tnsosfiles.com/).

The 2020 workbooks share the "All by Precinct" schema used by later years
(COUNTY, PRCTSEQ, PRECINCT, BALSEQID, ..., CANDGROUP, OFFICENAME, ELECTDATE,
ELECTTYPE, then up to 10 COLkHDG/RNAMEk/PARTYk/tallyk triples), with one
difference: the March workbook names the tally columns VOTES1..VOTES10 while
Aug/Nov name them PVTALLY1..PVTALLY10. The tally column prefix is detected
dynamically. ELECTTYPE distinguishes Republican/Democratic (Presidential
Preference) Primary / State General / Judicial Retention; every row of each
workbook belongs to that date's ballot, so no ELECTTYPE filtering is needed.

Conventions for the two PRIMARIES (standard OpenElections office names; Title-
case counties as printed; full party names; precinct names as printed with
internal whitespace collapsed so file_format passes; integer votes):
  "Candidates for President of the United States"  -> "Presidential Preference", "NA"
  "United States Senate"                           -> "U.S. Senate", "NA"
  "United States House of Representatives District N" -> "U.S. House", "N"
  "Tennessee Senate District N"                    -> "State Senate", "N"
  "Tennessee House of Representatives District N"  -> "State House", "N"
  "Delegate At-Large" / "Delegate District N"      -> "Delegate", "At-Large" / "N"
  "Supreme Court" / "Court of Appeals - ..." / "Court of Criminal Appeals - ..."
                                                   -> office as printed, "NA"
  "<judicial/DA/Public Defender/State Executive Committeeman/woman> ... District N"
                                                   -> office = text before "District N",
                                                       district "N" or "N (unexpired term)"
Judicial retention is non-partisan (PARTY column empty); the candidate is kept
verbatim ("Retain - <judge>" / "Replace - <judge>") with the party column empty.
The "State General" judicial/DA/Public Defender races on the Aug ballot DO carry
a party in the source (Republican/Democratic), which is kept. Write-in candidates
("Write-In - <name>") are kept verbatim with the party column empty when the
source party is empty.

The Nov 3 general COUNTY file deliberately matches the existing 2020 precinct
file's older convention (office "State Representative" rather than "State House",
empty district rather than "NA" for at-large offices, and float votes such as
"1554.0"), because it is derived by summing that precinct file. This keeps the
county + precinct pair internally consistent. (The Nov 3 ballot had no
constitutional amendments -- only President, U.S. Senate, U.S. House, State
Senate, State House -- so the existing precinct is complete.)

County totals are derived by summing precinct votes per
(county, office, district, party, candidate). Primary totals are spot-checked
against the official by-county PDFs; the general county totals are spot-checked
against "Nov 2020 General by County.pdf".

Requires `requests` (in the Pipfile), `openpyxl` (install separately, e.g.
`pip install openpyxl`) for the .xlsx workbooks, and `pdftotext` (poppler) only
if the by-county PDF spot-checks are run by hand.
"""

import csv
import os
import re
from collections import defaultdict

import requests

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("This parser requires openpyxl: pip install openpyxl") from exc

BASE = "https://sos-tn-gov-files.tnsosfiles.com"

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2020")

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party",
                   "candidate", "votes"]

USHOUSE_RE = re.compile(r"^United States House of Representatives District (\d+)\s*$")
TNSENATE_RE = re.compile(r"^Tennessee Senate District (\d+)\s*$")
TNHOUSE_RE = re.compile(r"^Tennessee House of Representatives District (\d+)\s*$")
DELEGATE_DIST_RE = re.compile(r"^Delegate District (\d+)\s*$")
DISTRICT_RE = re.compile(r"^(.+?)\s+District\s+(\d+(?:/\d+)?)\s*(\(unexpired term\))?\s*$")
WHITESPACE_RE = re.compile(r"\s+")


def clean(s):
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def norm_office(raw):
    """Map a source OFFICENAME to (office, district)."""
    s = raw.strip()
    if s == "Candidates for President of the United States":
        return ("Presidential Preference", "NA")
    if s in ("United States President", "President and Vice President of the United States"):
        return ("President", "NA")
    if s == "United States Senate":
        return ("U.S. Senate", "NA")
    m = USHOUSE_RE.match(s)
    if m:
        return ("U.S. House", m.group(1))
    m = TNSENATE_RE.match(s)
    if m:
        return ("State Senate", m.group(1))
    m = TNHOUSE_RE.match(s)
    if m:
        return ("State House", m.group(1))
    if s == "Delegate At-Large":
        return ("Delegate", "At-Large")
    m = DELEGATE_DIST_RE.match(s)
    if m:
        return ("Delegate", m.group(1))
    if s == "Supreme Court" or s.startswith("Court of Appeals") \
            or s.startswith("Court of Criminal Appeals"):
        return (s, "NA")
    m = DISTRICT_RE.match(s)
    if m:
        office = m.group(1).strip()
        dist = m.group(2) + (" (unexpired term)" if m.group(3) else "")
        return (office, dist)
    return (s, "NA")


def iter_xlsx(url):
    """Yield (county, precinct, office_raw, party, candidate, votes) per
    non-empty candidate slot across all CANDGROUP rows. Detects VOTES* (March)
    vs PVTALLY* (Aug/Nov) tally columns automatically."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(resp.content)
        f.flush()
        wb = openpyxl.load_workbook(f.name, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ci_county = header.index("COUNTY")
    ci_precinct = header.index("PRECINCT")
    ci_office = header.index("OFFICENAME")
    rname_cols, party_cols, tally_cols = [], [], []
    for i in range(1, 11):
        rn = f"RNAME{i}"
        py = f"PARTY{i}"
        if rn in header and py in header:
            tally = None
            for cand in (f"PVTALLY{i}", f"VOTES{i}"):
                if cand in header:
                    tally = cand
                    break
            if tally:
                rname_cols.append(header.index(rn))
                party_cols.append(header.index(py))
                tally_cols.append(header.index(tally))
    n = len(rname_cols)
    for r in rows:
        county = clean(r[ci_county] or "")
        precinct = clean(r[ci_precinct] or "")
        office_raw = clean(r[ci_office] or "")
        for k in range(n):
            name = r[rname_cols[k]]
            if not name:
                continue
            party = clean(r[party_cols[k]] or "")
            votes = r[tally_cols[k]] or 0
            yield (county, precinct, office_raw, party, clean(name), int(votes))


def build_primary(url):
    """Build county + precinct rows for one primary workbook."""
    precinct_rows = []
    county_sums = defaultdict(int)
    seen = {}
    dup_skipped = 0
    for county, precinct, office_raw, party, candidate, votes in iter_xlsx(url):
        office, district = norm_office(office_raw)
        key = (county, precinct, office, district, party, candidate)
        if key in seen:
            if seen[key] != votes:
                raise ValueError(f"conflicting duplicate precinct row: {key} "
                                 f"({seen[key]} vs {votes})")
            dup_skipped += 1
            continue
        seen[key] = votes
        precinct_rows.append([county, precinct, office, district, party, candidate, votes])
        county_sums[(county, office, district, party, candidate)] += votes
    if dup_skipped:
        print(f"  (skipped {dup_skipped} duplicate source rows)")
    county_rows = [[c, o, d, p, n, v] for (c, o, d, p, n), v in county_sums.items()]
    return county_rows, precinct_rows, county_sums


def build_general_county_from_precinct(precinct_path):
    """Derive the Nov 3 general COUNTY file by summing the existing precinct
    file, preserving its conventions (office names, empty districts, float
    votes) so county + precinct stay consistent."""
    sums = defaultdict(float)
    order = []
    seen = set()
    with open(precinct_path, newline="") as f:
        r = csv.reader(f)
        header = next(r)
        ci = {name: i for i, name in enumerate(header)}
        for row in r:
            county = row[ci["county"]]
            office = row[ci["office"]]
            district = row[ci["district"]]
            party = row[ci["party"]]
            candidate = row[ci["candidate"]]
            votes = float(row[ci["votes"]])
            key = (county, office, district, party, candidate)
            if key not in seen:
                seen.add(key)
                order.append(key)
            sums[key] += votes
    county_rows = [[k[0], k[1], k[2], k[3], k[4], sums[k]] for k in order]
    return county_rows


def sort_key_county(row):
    return (row[0], row[2], row[3], row[1], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  Wrote {path} ({len(rows)} rows)")


def main():
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)

    # Mar 3 Presidential Preference Primary.
    print("--- 20200303 primary (Mar 3, XLSX) ---")
    county_rows, precinct_rows, county_sums = build_primary(
        f"{BASE}/March2020Results.xlsx")
    county_rows.sort(key=sort_key_county)
    precinct_rows.sort(key=sort_key_precinct)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows, "
          f"{len(county_sums)} county/candidate combos")
    write_csv(os.path.join(out, "20200303__tn__primary__county.csv"),
              COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, "20200303__tn__primary__precinct.csv"),
              PRECINCT_HEADER, precinct_rows)

    # Aug 6 State Primary & General (full Aug 6 ballot).
    print("--- 20200806 primary (Aug 6, XLSX) ---")
    county_rows, precinct_rows, county_sums = build_primary(
        f"{BASE}/Aug2020PrecinctDetail.xlsx")
    county_rows.sort(key=sort_key_county)
    precinct_rows.sort(key=sort_key_precinct)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows, "
          f"{len(county_sums)} county/candidate combos")
    write_csv(os.path.join(out, "20200806__tn__primary__county.csv"),
              COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, "20200806__tn__primary__precinct.csv"),
              PRECINCT_HEADER, precinct_rows)

    # Nov 3 general, COUNTY ONLY -- derived from the existing precinct file so
    # it matches that file's conventions exactly (precinct file left untouched).
    print("--- 20201103 general (Nov 3, county from existing precinct) ---")
    precinct_path = os.path.join(out, "20201103__tn__general__precinct.csv")
    county_rows = build_general_county_from_precinct(precinct_path)
    county_rows.sort(key=sort_key_county)
    print(f"  county: {len(county_rows)} rows (derived from existing precinct)")
    write_csv(os.path.join(out, "20201103__tn__general__county.csv"),
              COUNTY_HEADER, county_rows)


if __name__ == "__main__":
    main()