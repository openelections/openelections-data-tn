"""
Parser for the 2010 Tennessee elections, producing OpenElections CSVs (county
+ precinct) for all three 2010 elections, sourced from the TN SoS results page
(https://sos.tn.gov/elections/results#2010):

  20100112__tn__special__general  -- Jan 12 HD83 special general (Shelby only).
        County-ONLY (the SoS source is a one-page county-total summary, no
        precinct breakdown): Mark White (R) 3236, Guthrie Castle (D) 1452,
        John Andreuccetti (I) 110.
  20100805__tn__primary  -- Aug 5 primary: Republican + Democratic primaries
        for Governor / U.S. House / State Senate / State House / State
        Executive Committeeman / Committeewoman, plus the non-partisan "State
        General" judicial contested races (Circuit/Criminal/Chancery Court,
        District Attorney General, Public Defender) and judicial retention
        (Supreme Court, Court of Appeals). Sourced from August2010.xlsx.
  20101102__tn__general  -- Nov 2 general: Governor, U.S. House, State Senate,
        State House, and one Constitutional Amendment (Yes/No). Sourced from
        November2010.xls.

The repo already had precinct-only CSVs for Aug 5 and Nov 2 (no county files, no
HD83 special) in a non-standard convention: UPPERCASE counties, float votes
(115.0), single-letter party (R/D), and -- for Aug 5 -- the State General
judicial/retention races dropped entirely. This parser REGENERATES them with
standard OpenElections conventions, ADDS the missing county files, and ADDS the
HD83 special-general county file.

Source workbooks (the SoS "All by Precinct" spreadsheets on
https://sos-prod.tnsosgovfiles.com/s3fs-public/document/): August2010.xlsx
(openpyxl) and November2010.xls (xlrd 2.0.2 -- reads .xls, NOT .xlsx). Shared
schema: COUNTY, (PRCTSEQ|Precinct Sequence), (PRECINCT|Precinct Name), Office
Seq, OfficeID, District, Candidate Group, OFFICENAME, ELECTDATE, ELECTTYPE,
then Col1/BName1/Tally1 ... Col10/BName10/Tally10 (the candidate number is in
ColN; BName is the name with a leading ". " leftover; TallyN is the precinct
vote). ELECTTYPE is "Republican Primary"/"Democratic Primary" (primaries) or
"State General" (Aug non-partisan judicial/retention + the entire Nov general).

Candidate Group handles >10 candidates: when a precinct-race has more than 10
candidates the source emits MULTIPLE rows for that precinct (Candidate Group 1
holding candidates 1-10 in BName1..BName10, Candidate Group 11 holding
candidates 11-20 in BName1..BName10 with Col=11..20). The parser simply emits
every non-empty BName cell across all rows, which captures every candidate
exactly once (the Nov 2 Governor race has 19 candidates incl. write-ins; the Aug
5 U.S. House Rep primaries in districts 3 and 5 overflow).

Conventions (standard OpenElections office names; title-case counties; integer
votes; full party names; bare legislative district number; precinct names
verbatim with internal whitespace collapsed):
  office  OFFICENAME ends with " District" + the number in the District column
            -> office = OFFICENAME minus " District" (then mapped to standard),
               district = District column value:
            "U.S. House of Representatives District" (Aug) /
            "U. S. House of Representatives District" (Nov) -> "U.S. House", <d>
            "Tennessee Senate District"            -> "State Senate", <d>
            "Tennessee House of Representatives District" -> "State House", <d>
            "State Executive Committeeman District" -> "State Executive
                                                         Committeeman", <d>
            "State Executive Committeewoman District" -> "State Executive
                                                         Committeewoman", <d>
            "Circuit/Criminal/Chancery Court ... District", "District Attorney
            General District", "Public Defender District" -> office as printed
                                                         (minus " District"), <d>
          OFFICENAME without " District":
            "Governor" -> "Governor", "NA"
            "Constitutional Amendment" -> "Constitutional Amendment", "NA"
            judicial retention (see below) -> <court>, "NA"
  retention  The 2010 source puts the judge in OFFICENAME ("Supreme Court -
            Sharon Gail Lee", "Court of Appeals - John W. McClarty - Eastern")
            and the BName is just "Retain"/"Replace". To match the repo's
            2012/2020/2022/2024 retention convention (office = the court,
            candidate = "Retain - <judge>"/"Replace - <judge>", party empty),
            the judge is moved into the candidate:
            "Supreme Court - Sharon Gail Lee" -> office "Supreme Court",
               candidate "Retain - Sharon Gail Lee" / "Replace - Sharon Gail Lee"
            "Court of Appeals - John W. McClarty - Eastern" -> office
               "Court of Appeals - Eastern Division" (matching the 2022 naming
               for the same Eastern-division seat), candidate
               "Retain - John W. McClarty" / "Replace - John W. McClarty"
  party   Primaries -> from ELECTTYPE (Republican / Democratic). The "State
            General" judicial contested races take party from each candidate's
            single-letter " - R"/" - D"/" - I" BName suffix (no parens, unlike
            2012's " - (R)"), stripped from the name and mapped to the full
            name; judicial candidates with no suffix (uncontested / non-partisan)
            get an empty party. Retention is non-partisan (party empty). The
            Nov 2 general takes party from the candidate's single-letter suffix.
  write-ins  "Write-in - <name>" -> normalized to "Write-In - <name>". In a
            primary the write-in candidate belongs to that primary's party
            (party = Republican/Democratic from ELECTTYPE, matching the 2018
            convention); in a general ("State General") the party is empty.
  "No Candidate Filed" is kept verbatim (party from ELECTTYPE), matching the
            2013/2018 special-primary convention.

Zero-vote rows ARE included (the source lists every candidate per precinct,
including "No Candidate Filed", retention, and write-ins at 0), matching the
dominant regenerated convention (2014/2016/2018/2024 all include 0-vote rows;
only 2012 and 2020 excluded them). County totals are the sum of the precinct
rows per (county, office, district, party, candidate).

County names are title-cased from the uppercase source (ANDERSON -> Anderson)
with fixes for the Mc/DeKalb counties (MCMINN -> McMinn, MCNAIRY -> McNairy,
DEKALB -> DeKalb).

The Jan 12 HD83 special-general county file is hardcoded from the one-page SoS
county-total summary PDF (Shelby only); there is no precinct source so no
precinct file is written.

Requires `requests` + `openpyxl` (in the Pipfile) and `xlrd==2.0.2` (reads
.xls; install separately, e.g. `pip install xlrd==2.0.2`).
"""

import csv
import io
import os
import re
import tempfile
from collections import defaultdict

import openpyxl
import requests
import xlrd

BASE_XLSX = "https://sos-prod.tnsosgovfiles.com/s3fs-public/document"
AUG_URL = f"{BASE_XLSX}/August2010.xlsx"
NOV_URL = f"{BASE_XLSX}/November2010.xls"

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2010")

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party",
                   "candidate", "votes"]

WHITESPACE_RE = re.compile(r"\s+")
# Single-letter party suffix, NO parens ("Bill Haslam - R"); only treated as a
# party when the letter is a known party (R/D/C/G/I/L) -- see parse_candidate.
PARTY_SUFFIX_RE = re.compile(r"^(.+?)\s+-\s+([A-Z])\s*$")
PARTY_LETTER = {"R": "Republican", "D": "Democratic", "C": "Constitution",
                "G": "Green", "I": "Independent", "L": "Libertarian"}

OFFICE_MAP = {
    "Governor": "Governor",
    "U.S. House of Representatives": "U.S. House",
    "U. S. House of Representatives": "U.S. House",
    "Tennessee Senate": "State Senate",
    "Tennessee House of Representatives": "State House",
    "State Executive Committeeman": "State Executive Committeeman",
    "State Executive Committeewoman": "State Executive Committeewoman",
    "Constitutional Amendment": "Constitutional Amendment",
}

# Documented source-typo fix: one Shelby HD90 (Aug Dem primary) precinct
# misspells the incumbent "John J. Deberry, Jr." as "John J  Deberry, Jr."
# (no period after the middle initial, double space). The authoritative SoS
# by-county PDF aggregates him as a single "John J. Deberry, Jr." (6175 votes);
# without this alias he splits into two candidates (5764 + 411) in the county
# file. Applied to the cleaned candidate name.
CANDIDATE_ALIASES = {
    "John J Deberry, Jr.": "John J. Deberry, Jr.",
}

# 2010 judicial-retention OFFICENAME -> (court office, judge). The judge is in
# the OFFICENAME (and the BName is just "Retain"/"Replace"); we move the judge
# into the candidate to match the 2012/2020/2022/2024 retention convention.
RETENTION = {
    "Supreme Court - Sharon Gail Lee":
        ("Supreme Court", "Sharon Gail Lee"),
    "Court of Appeals - John W. McClarty - Eastern":
        ("Court of Appeals - Eastern Division", "John W. McClarty"),
}


def clean(s):
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def norm_county(c):
    s = clean(c)
    if not s:
        return s
    parts = s.title().split()
    fixed = []
    for p in parts:
        low = p.lower()
        if low == "dekalb":
            p = "DeKalb"
        elif low.startswith("mc") and len(p) > 2:
            p = "Mc" + p[2].upper() + p[3:]
        fixed.append(p)
    return " ".join(fixed)


def norm_district(d):
    if d is None:
        return "NA"
    s = clean(d)
    if s == "" or s.lower() == "none":
        return "NA"
    # numeric -> bare number (xlrd may give "3.0" for a numeric district)
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


def primary_party(etype):
    et = clean(etype)
    if "Republican" in et:
        return "Republican"
    if "Democratic" in et:
        return "Democratic"
    return ""


def norm_office(raw_office, district_col):
    """Return (office, district, judge_or_None) from OFFICENAME + District col.
    judge_or_None is set for retention offices (the candidate is then built as
    "Retain/Replace - <judge>")."""
    raw = clean(raw_office)
    if raw in RETENTION:
        court, judge = RETENTION[raw]
        return court, "NA", judge
    if raw.endswith(" District"):
        base = raw[: -len(" District")].strip()
    else:
        base = raw
    office = OFFICE_MAP.get(base, base)
    return office, norm_district(district_col), None


def parse_candidate(bname, etype, judge=None):
    """Return (name, party) from a BName cell + ELECTTYPE.

    For retention (judge is not None) the candidate is "<Retain|Replace> -
    <judge>" with an empty party. Otherwise write-ins are normalized to
    "Write-In - <rest>" (party = the primary party, or "" for a general); a
    single-letter party suffix is stripped and mapped to the full name; bare
    names take party from the ELECTTYPE (primary) or "" (State General)."""
    name = clean(bname)
    # strip the leading ". " leftover (candidate number is in the Col field)
    if name.startswith(". "):
        name = name[2:].strip()
    elif name.startswith("."):
        name = name[1:].strip()
    if not name:
        return "", ""
    if judge is not None:
        return f"{name} - {judge}", ""
    if name.lower().startswith("write-in"):
        rest = name[len("write-in"):].lstrip()
        if rest[:1] == "-":
            rest = rest[1:].lstrip()
        # a write-in may itself carry a party suffix (e.g.
        # "Write-in - Colonel Gean Billingsley - D"); strip it into the party
        # column like every other "- X" candidate. Otherwise a primary
        # write-in takes the primary's party; a general write-in is empty.
        party = primary_party(etype)
        m = PARTY_SUFFIX_RE.match(rest)
        if m and m.group(2) in PARTY_LETTER:
            rest = clean(m.group(1))
            party = PARTY_LETTER[m.group(2)]
        return f"Write-In - {rest}".strip(), party
    m = PARTY_SUFFIX_RE.match(name)
    if m and m.group(2) in PARTY_LETTER:
        return CANDIDATE_ALIASES.get(clean(m.group(1)), clean(m.group(1))), \
            PARTY_LETTER[m.group(2)]
    return CANDIDATE_ALIASES.get(name, name), primary_party(etype)


def _column_index(header, *candidates):
    """Return the index of the first header name in `candidates` that appears
    in `header` (handles the Aug 'PRECINCT' vs Nov 'Precinct Name' rename)."""
    for c in candidates:
        if c in header:
            return header.index(c)
    raise KeyError(f"none of {candidates} found in header")


def _iter_workbook(url):
    """Yield (county, precinct, office_raw, district_col, etype, [bname cells],
    [tally cells]) per source row. Handles .xlsx (openpyxl) and .xls (xlrd)."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    data = resp.content
    if url.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=data)
        ws = wb.sheet_by_index(0)
        nrows = ws.nrows
        def rowvals(i):
            return [ws.cell_value(i, j) for j in range(ws.ncols)]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                    data_only=True)
        ws = wb[wb.sheetnames[0]]
        allrows = list(ws.iter_rows(values_only=True))
        nrows = len(allrows)
        def rowvals(i):
            return list(allrows[i])
    header = [clean(c) if c is not None else "" for c in rowvals(0)]
    county_i = _column_index(header, "COUNTY")
    precinct_i = _column_index(header, "PRECINCT", "Precinct Name")
    office_i = _column_index(header, "OFFICENAME")
    etype_i = _column_index(header, "ELECTTYPE")
    dist_i = _column_index(header, "District")
    bname_cols = [i for i, c in enumerate(header) if c.startswith("BName")]
    tally_cols = [i for i, c in enumerate(header) if c.startswith("Tally")]
    for i in range(1, nrows):
        r = rowvals(i)
        yield (r[county_i], r[precinct_i], r[office_i], r[dist_i],
               r[etype_i],
               [r[c] for c in bname_cols],
               [r[c] for c in tally_cols])


def to_int(v):
    if v is None:
        return 0
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if s == "":
        return 0
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return 0


def load_workbook(url):
    """Parse one All-by-Precinct workbook. Returns (precinct_rows, county_rows).
    Every non-empty BName cell is emitted (this captures the >10-candidate
    overflow rows exactly once). Zero-vote rows are included."""
    precinct_rows = []
    county_sum = defaultdict(int)
    for (county, precinct, office_raw, district_col, etype,
         bnames, tallies) in _iter_workbook(url):
        if office_raw is None or not str(office_raw).strip():
            continue
        county = norm_county(county)
        precinct = clean(precinct)
        office, district, judge = norm_office(office_raw, district_col)
        for k, bn in enumerate(bnames):
            if bn is None or not str(bn).strip() or str(bn).strip() == ".":
                continue
            name, party = parse_candidate(bn, etype, judge)
            if not name:
                continue
            votes = to_int(tallies[k])
            precinct_rows.append([county, precinct, office, district, party,
                                  name, votes])
            county_sum[(county, office, district, party, name)] += votes
    county_rows = [[c, o, d, p, n, v]
                   for (c, o, d, p, n), v in county_sum.items()]
    return precinct_rows, county_rows


def sort_key_county(row):
    return (row[0], row[1], row[2], row[3], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  Wrote {path} ({len(rows)} rows)")


def build_hd83():
    """Hardcode the Jan 12 HD83 special-general county file from the one-page
    SoS county-total summary (Shelby only; no precinct source)."""
    rows = [
        ["Shelby", "State House", "83", "Republican", "Mark White", 3236],
        ["Shelby", "State House", "83", "Democratic", "Guthrie Castle", 1452],
        ["Shelby", "State House", "83", "Independent", "John Andreuccetti",
         110],
    ]
    rows.sort(key=sort_key_county)
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)
    write_csv(os.path.join(out, "20100112__tn__special__general__county.csv"),
              COUNTY_HEADER, rows)


def build(name, url):
    print(f"--- {name} ---")
    precinct_rows, county_rows = load_workbook(url)
    precinct_rows.sort(key=sort_key_precinct)
    county_rows.sort(key=sort_key_county)
    print(f"  county: {len(county_rows)} rows, precinct: {len(precinct_rows)} rows")
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)
    write_csv(os.path.join(out, f"{name}__county.csv"), COUNTY_HEADER, county_rows)
    write_csv(os.path.join(out, f"{name}__precinct.csv"), PRECINCT_HEADER,
              precinct_rows)


def main():
    print("--- 20100112__tn__special__general (HD83) ---")
    build_hd83()
    build("20100805__tn__primary", AUG_URL)
    build("20101102__tn__general", NOV_URL)


if __name__ == "__main__":
    main()