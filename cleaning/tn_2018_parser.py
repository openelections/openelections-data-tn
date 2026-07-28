"""
Parser for the 2018 Tennessee elections, producing OpenElections CSVs (county +
precinct) for both 2018 primaries, the two SD14 special elections, and the Nov 6
general -- all with standard OpenElections office names. Per the user's choice,
the two elections that already had precinct CSVs in the repo (Aug 2 primary, Nov
6 general) are REGENERATED from the XLSX with standard names (overwriting those
existing precinct files) and their county files added.

  20180125__tn__special__primary  -- Jan 25 Senate District 14 special primary
                                      (Republican + Democratic primaries merged)
  20180313__tn__special__general  -- Mar 13 Senate District 14 special general
  20180501__tn__primary           -- May 1 State Judicial Primary (Republican +
                                      Democratic primaries merged)
  20180802__tn__primary           -- Aug 2 State Primary & General: the full
                                      Aug 2 ballot (Democratic + Republican
                                      primaries for Governor/U.S. Senate/U.S.
                                      House/State Senate/State House/State
                                      Executive Committeeman/woman; the judicial
                                      "State General" seats; judicial retention)
  20181106__tn__general           -- Nov 6 general (Governor, U.S. Senate,
                                      U.S. House, State Senate, State House)

Sources (https://sos.tn.gov/elections/results#2018; WebFetch 403s, use curl):
  Jan 25 special primary -- by-precinct PDFs on the OLDER s3 bucket
      https://sos-tn-gov-files.s3.amazonaws.com/
      (TN Senate 14 {Rep,Dem} Precinct.pdf)
  Mar 13 special general -- by-precinct PDF on sos-tn-gov-files.tnsosfiles.com
      (TN Senate 14 Precincts.pdf)
  May 1 judicial primary -- by-precinct PDFs (multi-race per PDF) on
      sos-tn-gov-files.tnsosfiles.com
      (State Offices {Republican,Democratic} Primary Precinct Totals.pdf)
  Aug 2  -- 180802_ResultsbyPrecinct.xlsx  (NO CANDGROUP column; one row per
      precinct/office with up to 10 inline RNAMEk/PARTYk/PVTALLYk triples)
  Nov 6  -- Nov 2018 General results.xlsx  (HAS CANDGROUP; the standard
      All-by-Precinct schema used by later years)
  Both XLSX on https://sos-tn-gov-files.tnsosfiles.com/.

The by-precinct PDFs (specials + May 1) use the multi-county "Layout B" format:
  Banner:     State of Tennessee
  Date:       <Month Day, Year>
  Section:    Special Republican Primary | Special Democratic Primary |
              Special State General | Republican Primary | Democratic Primary |
              State General   (2021 used "Special State General Election")
  Office:     Tennessee Senate District N [(unexpired term)]
              Tennessee House of Representatives District N [(unexpired term)]
              <judicial/DA/Public Defender/State Exec Committeeman> ... District N
              [(unexpired term)]
  Candidates: N. Name            (primaries; party from the section header)
              N. Name - Party    (generals; party from the suffix)
  Column hdr: 1 2 3 ... N        (present only when >1 candidate)
  <County> County
   Precincts:
    <precinct name>  v1 v2 ... vN      (votes may contain commas, e.g. 1,512)
   County Totals:   v1 v2 ... vN
  ...
   DISTRICT TOTALS  v1 v2 ... vN
A single PDF may contain MANY races (May 1 judicial primary): each race repeats
the banner/section/office/candidates header, then its county/precinct blocks.
Multi-page races repeat the header on each page; the parser is a state machine
that re-reads the (office, candidates) on each header and accumulates precinct
rows, so repeated headers for the same race simply yield more precinct rows.

Conventions (standard OpenElections office names; integer votes; full party
names; precinct names verbatim with internal whitespace collapsed):
  "United States Senate"                            -> "U.S. Senate", "NA"
  "United States House of Representatives District N" -> "U.S. House", "N"
  "Tennessee Senate District N [(unexpired term)]"  -> "State Senate", "N" (bare)
  "Tennessee House of Representatives District N [(unexpired term)]"
                                                     -> "State House", "N" (bare)
  "Governor"                                        -> "Governor", "NA"
  "<judicial/DA/Public Defender/State Exec Committeeman/woman> ... District N
       [(unexpired term)]"                          -> office = text before
       "District N", district "N" or "N (unexpired term)"
Legislative offices use a BARE district number -- the "(unexpired term)" suffix
is for judicial/executive races, not legislative ones, so it is stripped for
Senate/House (matching the 2013/2015/2019/2021/2023 special convention) but
preserved for judicial/DA/Public Defender/State Exec Committeeman races.
Judicial retention is non-partisan (candidate kept verbatim
"Retain - <judge>"/"Replace - <judge>", party empty); the Aug "State General"
judicial/DA/Public Defender races carry a party in the source, which is kept.

County totals are derived by summing precinct votes per (county, office,
district, party, candidate). For the PDF elections they are asserted against
each precinct PDF's own "County Totals:" lines; for the XLSX elections they are
spot-checked against the official by-county PDFs.

Requires `requests` (in the Pipfile), `openpyxl` (install separately, e.g.
`pip install openpyxl`) for the .xlsx workbooks, and `pdftotext` (poppler) for
the by-precinct PDFs.
"""

import csv
import os
import re
import subprocess
import tempfile
from collections import defaultdict

import requests

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("This parser requires openpyxl: pip install openpyxl") from exc

BASE = "https://sos-tn-gov-files.tnsosfiles.com"
BASE_S3 = "https://sos-tn-gov-files.s3.amazonaws.com"

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "2018")

COUNTY_HEADER = ["county", "office", "district", "party", "candidate", "votes"]
PRECINCT_HEADER = ["county", "precinct", "office", "district", "party",
                   "candidate", "votes"]

SECTION_RE = re.compile(
    r"^(Special Republican Primary|Special Democratic Primary|"
    r"Special State General Election|Special State General|"
    r"Republican Primary|Democratic Primary|State General)\s*$")
OFFICE_SENATE_RE = re.compile(
    r"^Tennessee Senate District (\d+)(?:\s*\(unexpired term\))?\s*$")
OFFICE_HOUSE_RE = re.compile(
    r"^Tennessee House of Representatives District (\d+)"
    r"(?:\s*\(unexpired term\))?\s*$")
USHOUSE_RE = re.compile(
    r"^United States House of Representatives District (\d+)\s*$")
DISTRICT_RE = re.compile(
    r"^(.+?)\s+District\s+(\d+(?:/\d+)?)\s*(\(unexpired term\))?\s*$")
CAND_RE = re.compile(r"^\s*(\d+)\.\s+(.+?)\s*$")
COUNTY_RE = re.compile(r"^([A-Z][A-Za-z .]+) County\s*$")
PARTY_SUFFIX_RE = re.compile(
    r"^(.+?)\s+-\s+(Republican|Democratic|Independent|Libertarian|Green|"
    r"Constitution)\s*$")
VOTE_RE = re.compile(r"^\d[\d,]*$")
ALLDIGITS_RE = re.compile(r"^\d+$")
WHITESPACE_RE = re.compile(r"\s+")


def clean(s):
    return WHITESPACE_RE.sub(" ", str(s)).strip()


def to_int(tok):
    return int(tok.replace(",", ""))


def norm_office(raw):
    """Map a source office string (XLSX OFFICENAME or PDF office line) to
    (office, district). Legislative offices get a BARE district (the
    "(unexpired term)" suffix is stripped); judicial/DA/etc keep the suffix."""
    s = raw.strip()
    if s == "United States Senate":
        return ("U.S. Senate", "NA")
    if s == "Governor":
        return ("Governor", "NA")
    m = USHOUSE_RE.match(s)
    if m:
        return ("U.S. House", m.group(1))
    m = OFFICE_SENATE_RE.match(s)
    if m:
        return ("State Senate", m.group(1))
    m = OFFICE_HOUSE_RE.match(s)
    if m:
        return ("State House", m.group(1))
    m = DISTRICT_RE.match(s)
    if m:
        office = m.group(1).strip()
        dist = m.group(2) + (" (unexpired term)" if m.group(3) else "")
        return (office, dist)
    return (s, "NA")


def split_candidate(raw, section):
    """Return (name, party). Primaries take party from the section header;
    generals take it from the candidate's ` - Party` suffix."""
    if section in ("Special Republican Primary", "Republican Primary"):
        return (raw.strip(), "Republican")
    if section in ("Special Democratic Primary", "Democratic Primary"):
        return (raw.strip(), "Democratic")
    m = PARTY_SUFFIX_RE.match(raw.strip())
    if m:
        return (m.group(1).strip(), m.group(2))
    return (raw.strip(), "")


# ---------------------------------------------------------------------------
# PDF parsing (specials + May 1 judicial primary)
# ---------------------------------------------------------------------------

def pdftotext_layout(pdf_bytes):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        path = f.name
    try:
        out = subprocess.run(
            ["pdftotext", "-layout", path, "-"],
            check=True, capture_output=True, text=True)
        return out.stdout
    finally:
        os.unlink(path)


def parse_pdf(text):
    """Parse a (possibly multi-race, multi-page) Layout B by-precinct PDF.
    Returns (precinct_rows, county_totals) where precinct_rows are
    [county, precinct, office, district, party, candidate, votes] and
    county_totals maps (county, office, district, party, candidate) -> votes
    from the PDFs' own "County Totals:" lines (for cross-checking)."""
    precinct_rows = []
    county_totals = {}
    section = None
    office = None
    district = None
    candidates = []          # list of (name, party)
    expect_office = False
    collecting_candidates = False
    in_precincts = False
    current_county = None
    for line in text.splitlines():
        s = line.strip()
        # Page-break sentinels: the footer ("<date> Page N of M") and the
        # next-page banner ("State of Tennessee") appear between a page's last
        # precinct row and the reprinted race header. End any in-progress
        # precinct block so the date/banner lines that follow aren't mis-parsed
        # as precinct rows (e.g. "May 1, 2018" -> precinct "May" w/ votes
        # [1, 2018], since "1," and "2018" both look like vote tokens).
        if "Page " in s and " of " in s:
            in_precincts = False
            continue
        if s == "State of Tennessee":
            in_precincts = False
            continue
        m = SECTION_RE.match(s)
        if m:
            section = s
            office = None
            district = None
            candidates = []
            expect_office = True
            collecting_candidates = False
            in_precincts = False
            current_county = None
            continue
        if expect_office:
            if not s:
                continue
            office, district = norm_office(s)
            expect_office = False
            collecting_candidates = True
            continue
        if collecting_candidates:
            if not s:
                continue
            cm = CAND_RE.match(line)
            if cm:
                candidates.append(split_candidate(cm.group(2).strip(), section))
                continue
            toks = s.split()
            if toks and all(ALLDIGITS_RE.match(t) for t in toks):
                # column-header line ("1 2 3 4")
                continue
            # First non-candidate, non-column-header line starts the body.
            collecting_candidates = False
            # fall through to body handling for this line
        # body handling
        if not s:
            continue
        cm = COUNTY_RE.match(line)
        if cm:
            current_county = cm.group(1).strip()
            in_precincts = False
            continue
        if s == "Precincts:":
            in_precincts = True
            continue
        if s.startswith("County Totals:"):
            nums = s[len("County Totals:"):].split()
            if len(nums) == len(candidates):
                for k, (name, party) in enumerate(candidates):
                    county_totals[(current_county, office, district, party,
                                   name)] = to_int(nums[k])
            in_precincts = False
            continue
        if s.startswith("DISTRICT TOTALS"):
            in_precincts = False
            continue
        if in_precincts and current_county and candidates:
            n = len(candidates)
            tokens = line.split()
            if len(tokens) < n + 1:
                continue
            vote_tokens = tokens[-n:]
            if not all(VOTE_RE.match(t) for t in vote_tokens):
                continue
            precinct = clean(" ".join(tokens[:-n]))
            for k, (name, party) in enumerate(candidates):
                precinct_rows.append([current_county, precinct, office,
                                      district, party, name, to_int(vote_tokens[k])])
    return precinct_rows, county_totals


def build_pdf(election):
    """Build county + precinct rows for one PDF-based election (which may
    combine several PDFs, e.g. a primary's Rep+Dem PDFs). Cross-checks derived
    county totals against the PDFs' own County Totals lines."""
    precinct_rows = []
    expected = {}
    for pdf_url in election["pdfs"]:
        resp = requests.get(pdf_url, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        text = pdftotext_layout(resp.content)
        prows, ctotals = parse_pdf(text)
        precinct_rows.extend(prows)
        for k, v in ctotals.items():
            expected[k] = v
    derived = defaultdict(int)
    for r in precinct_rows:
        county, precinct, office, district, party, name, votes = r
        derived[(county, office, district, party, name)] += votes
    mismatches = []
    for k, v in derived.items():
        if k in expected and expected[k] != v:
            mismatches.append((k, v, expected[k]))
    for k, v in expected.items():
        if k not in derived:
            mismatches.append((k, "MISSING", v))
    if mismatches:
        detail = "\n  ".join(
            f"{k}: derived={dv} expected={ev}" for k, dv, ev in mismatches[:20])
        raise AssertionError(f"county-total mismatch in {election['name']}:\n  "
                            f"{detail}")
    county_rows = [[k[0], k[1], k[2], k[3], k[4], v]
                   for k, v in derived.items()]
    return county_rows, precinct_rows, derived


# ---------------------------------------------------------------------------
# XLSX parsing (Aug 2 primary, Nov 6 general)
# ---------------------------------------------------------------------------

def iter_xlsx(url):
    """Yield (county, precinct, office_raw, party, candidate, votes) per
    non-empty candidate slot. Handles both the CANDGROUP schema (Nov 6) and the
    no-CANDGROUP inline schema (Aug 2) -- both put candidates in RNAMEk/PARTYk/
    PVTALLYk columns, which is all this iterates."""
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(resp.content)
        f.flush()
        wb = openpyxl.load_workbook(f.name, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    ci_county = header.index("COUNTY")
    ci_precinct = header.index("PRECINCT")
    ci_office = header.index("OFFICENAME")
    rname_cols, party_cols, tally_cols = [], [], []
    for i in range(1, 11):
        rn, py = f"RNAME{i}", f"PARTY{i}"
        if rn in header and py in header:
            tally = None
            for cand in (f"PVTALLY{i}", f"VOTES{i}"):
                if cand in header:
                    tally = cand
                    break
            if tally:
                rname_cols.append(header.index(rn))
                party_cols.append(header.index(py))
                tally_cols.append(header.index(tally))
    n = len(rname_cols)
    for r in rows:
        county = clean(r[ci_county] or "")
        precinct = clean(r[ci_precinct] or "")
        office_raw = clean(r[ci_office] or "")
        for k in range(n):
            name = r[rname_cols[k]]
            if not name:
                continue
            party = clean(r[party_cols[k]] or "")
            votes = r[tally_cols[k]] or 0
            yield (county, precinct, office_raw, party, clean(name), int(votes))


def build_xlsx(url):
    """Build county + precinct rows for one XLSX election (keep-first dedup,
    raise on conflicting duplicate rows)."""
    precinct_rows = []
    county_sums = defaultdict(int)
    seen = {}
    dup_skipped = 0
    for county, precinct, office_raw, party, candidate, votes in iter_xlsx(url):
        office, district = norm_office(office_raw)
        key = (county, precinct, office, district, party, candidate)
        if key in seen:
            if seen[key] != votes:
                raise ValueError(f"conflicting duplicate precinct row: {key} "
                                 f"({seen[key]} vs {votes})")
            dup_skipped += 1
            continue
        seen[key] = votes
        precinct_rows.append([county, precinct, office, district, party,
                              candidate, votes])
        county_sums[(county, office, district, party, candidate)] += votes
    if dup_skipped:
        print(f"  (skipped {dup_skipped} duplicate source rows)")
    county_rows = [[c, o, d, p, n, v]
                   for (c, o, d, p, n), v in county_sums.items()]
    return county_rows, precinct_rows, county_sums


# ---------------------------------------------------------------------------
# output helpers
# ---------------------------------------------------------------------------

def sort_key_county(row):
    return (row[0], row[2], row[3], row[1], row[4])


def sort_key_precinct(row):
    return (row[0], row[1], row[2], row[3], row[4], row[5])


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  Wrote {path} ({len(rows)} rows)")


ELECTIONS = [
    {
        "name": "20180125__tn__special__primary",
        "kind": "pdf",
        "pdfs": [
            f"{BASE_S3}/TN%20Senate%2014%20Rep%20Precinct.pdf",
            f"{BASE_S3}/TN%20Senate%2014%20Dem%20Precinct.pdf",
        ],
    },
    {
        "name": "20180313__tn__special__general",
        "kind": "pdf",
        "pdfs": [
            f"{BASE}/TN%20Senate%2014%20Precincts.pdf",
        ],
    },
    {
        "name": "20180501__tn__primary",
        "kind": "pdf",
        "pdfs": [
            f"{BASE}/State%20Offices%20Republican%20Primary%20Precinct%20Totals.pdf",
            f"{BASE}/State%20Offices%20Democratic%20Primary%20Precinct%20Totals.pdf",
        ],
    },
    {
        "name": "20180802__tn__primary",
        "kind": "xlsx",
        "url": f"{BASE}/180802_ResultsbyPrecinct.xlsx",
    },
    {
        "name": "20181106__tn__general",
        "kind": "xlsx",
        "url": f"{BASE}/Nov%202018%20General%20results.xlsx",
    },
]


def main():
    out = os.path.abspath(OUT_DIR)
    os.makedirs(out, exist_ok=True)
    for election in ELECTIONS:
        print(f"--- {election['name']} ({election['kind']}) ---")
        if election["kind"] == "pdf":
            county_rows, precinct_rows, _ = build_pdf(election)
        else:
            county_rows, precinct_rows, _ = build_xlsx(election["url"])
        county_rows.sort(key=sort_key_county)
        precinct_rows.sort(key=sort_key_precinct)
        print(f"  county: {len(county_rows)} rows, precinct: "
              f"{len(precinct_rows)} rows")
        # The election name already carries the right infix
        # (__special__primary/__special__general for the SD14 specials, plain
        # __primary for May 1/Aug 2, __general for Nov 6).
        write_csv(os.path.join(out, f"{election['name']}__county.csv"),
                  COUNTY_HEADER, county_rows)
        write_csv(os.path.join(out, f"{election['name']}__precinct.csv"),
                  PRECINCT_HEADER, precinct_rows)


if __name__ == "__main__":
    main()