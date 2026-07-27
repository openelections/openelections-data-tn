"""
Parser for the 2025 Tennessee U.S. House District 7 special elections
(unexpired term), covering both the October 7, 2025 special primary and the
December 2, 2025 special general.

Source (TN SoS "All by Precinct" spreadsheets, which list every precinct's
votes for every candidate; grand totals were cross-checked against the official
by-county PDFs and match exactly):
  primary:  https://sos-prod.tnsosgovfiles.com/s3fs-public/document/20251007AllbyPrecinct.xlsx
  general:  https://sos-prod.tnsosgovfiles.com/s3fs-public/document/20251202AllbyPrecinct.xlsx

Outputs (in 2025/):
  20251007__tn__special__primary__county.csv
  20251007__tn__special__primary__precinct.csv
  20251202__tn__special__general__county.csv
  20251202__tn__special__general__precinct.csv

The XLSX is a wide layout: one row per (precinct, candidate-group) with up to 10
candidates per row in RNAME*/PARTY*/PVTALLY* columns. A precinct may span
multiple rows when a party fields more than 10 candidates (the Republican
primary had 11). County-level totals are derived by summing the precinct votes
per county, which matches the official by-county PDFs exactly. The
"All County Precinct" bucket (absentee/early votes reported at the county level)
is kept as a precinct, matching the repo's 2022 convention, so precinct sums
reconcile to county totals.

Conventions (matching the repo's canonical files):
  county:   Title case, e.g. "Humphreys" (the primary source misspells it
            "Humhreys" and the general source uppercases it; both are normalized)
  office:   "U.S. House"
  district: "7"
  party:    full names ("Democratic", "Republican", "Independent")
  precinct: as printed in the source (e.g. "1 Holladay", "All County Precinct")

Requires `requests` (in the Pipfile) and `openpyxl` (a third-party package;
install separately, e.g. `pip install openpyxl`) to read the .xlsx workbook,
which is far more reliable than parsing the multi-county PDFs.
"""

import csv
import os
import re
from collections import defaultdict

import requests

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("This parser requires openpyxl: pipenv install openpyxl") from exc

BASE = "https://sos-prod.tnsosgovfiles.com/s3fs-public/document"
PRIMARY_URL = f"{BASE}/20251007AllbyPrecinct.xlsx"
GENERAL_URL = f"{BASE}/20251202AllbyPrecinct.xlsx"

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "2025")

OFFICE = "U.S. House"
DISTRICT = "7"

# Canonical Title-case county names, keyed by lowercased source spelling so the
# primary's "Humhreys" typo and the general's all-caps names both normalize.
COUNTY_CANON = {
    "benton": "Benton", "cheatham": "Cheatham", "davidson": "Davidson",
    "decatur": "Decatur", "dickson": "Dickson", "hickman": "Hickman",
    "houston": "Houston", "humphreys": "Humphreys", "humhreys": "Humphreys",
    "montgomery": "Montgomery", "perry": "Perry", "robertson": "Robertson",
    "stewart": "Stewart", "wayne": "Wayne", "williamson": "Williamson",
}

PARTY_ORDER = {"Democratic": 0, "Republican": 1, "Independent": 2}


def norm_county(raw):
    return COUNTY_CANON.get(str(raw).strip().lower(), str(raw).strip().title())


def fetch_workbook(url):
    resp = requests.get(url)
    resp.raise_for_status()
    # openpyxl reads from a file-like object via load_workbook; write bytes to a
    # temp file to avoid seeking issues with some stream wrappers.
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(resp.content)
        f.flush()
        return openpyxl.load_workbook(f.name, read_only=True, data_only=True)


def iter_precinct_rows(url):
    """Yield (county, precinct, party, candidate, votes) for every candidate."""
    wb = fetch_workbook(url)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    rname_cols = [header.index(f"RNAME{i}") for i in range(1, 11)
                  if f"RNAME{i}" in header]
    party_cols = [header.index(f"PARTY{i}") for i in range(1, 11)
                  if f"PARTY{i}" in header]
    tally_cols = [header.index(f"PVTALLY{i}") for i in range(1, 11)
                  if f"PVTALLY{i}" in header]
    n = min(len(rname_cols), len(party_cols), len(tally_cols))
    ci_county = header.index("COUNTY")
    ci_precinct = header.index("PRECINCT")
    for r in rows:
        county = norm_county(r[ci_county])
        precinct = str(r[ci_precinct]).strip()
        for k in range(n):
            name = r[rname_cols[k]]
            if not name:
                continue
            party = str(r[party_cols[k]] or "").strip()
            votes = r[tally_cols[k]] or 0
            yield county, precinct, party, str(name).strip(), int(votes)


def sort_key_county(row):
    # row: [county, office, district, party, candidate, votes]
    return (row[0], PARTY_ORDER.get(row[3], 99), row[4])


def sort_key_precinct(row):
    # row: [county, precinct, office, district, party, candidate, votes]
    return (row[0], row[1], PARTY_ORDER.get(row[4], 99), row[5])


def build(url, county_path, precinct_path):
    precinct_rows = []
    county_sums = defaultdict(int)  # (county, party, candidate) -> votes
    for county, precinct, party, candidate, votes in iter_precinct_rows(url):
        precinct_rows.append(
            [county, precinct, OFFICE, DISTRICT, party, candidate, votes]
        )
        county_sums[(county, party, candidate)] += votes

    county_rows = [
        [county, OFFICE, DISTRICT, party, candidate, votes]
        for (county, party, candidate), votes in county_sums.items()
    ]
    county_rows.sort(key=sort_key_county)
    precinct_rows.sort(key=sort_key_precinct)

    write_csv(
        county_path,
        ["county", "office", "district", "party", "candidate", "votes"],
        county_rows,
    )
    write_csv(
        precinct_path,
        ["county", "precinct", "office", "district", "party", "candidate", "votes"],
        precinct_rows,
    )

    # Sanity check: precinct sums must equal county totals.
    ps = defaultdict(int)
    for r in precinct_rows:
        ps[(r[0], r[4], r[5])] += r[6]
    for key, v in county_sums.items():
        assert ps[key] == v, f"mismatch {key}: {ps[key]} != {v}"
    return county_rows, precinct_rows


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"Wrote {path} ({len(rows)} rows)")


def main():
    county_dir = os.path.abspath(OUT_DIR)
    os.makedirs(county_dir, exist_ok=True)

    build(
        PRIMARY_URL,
        os.path.join(county_dir, "20251007__tn__special__primary__county.csv"),
        os.path.join(county_dir, "20251007__tn__special__primary__precinct.csv"),
    )
    build(
        GENERAL_URL,
        os.path.join(county_dir, "20251202__tn__special__general__county.csv"),
        os.path.join(county_dir, "20251202__tn__special__general__precinct.csv"),
    )


if __name__ == "__main__":
    main()